from openpyxl import load_workbook
from io import BytesIO

from django.contrib.auth.models import AnonymousUser, User, Group, Permission
from django.test import TestCase, RequestFactory
from django.core.urlresolvers import reverse
from django.contrib.messages.storage.fallback import FallbackStorage
from django.http import Http404
from django.conf import settings

from grid.views.deal import ChangeDealView, DealDetailView, DeleteDealView, RecoverDealView, AddDealView
from grid.views.export import ExportView
from editor.views import ManageAddsView, ManageUpdatesView, ManageDeletesView, ManageForUserView, \
    ApproveActivityChangeView, ApproveActivityDeleteView, LogAddedView, LogModifiedView, LogDeletedView
from landmatrix.models import HistoricalActivity
from api.elasticsearch import es_search

class BaseTestDeal(TestCase):
    fixtures = [
        'countries_and_regions',
        'users_and_groups',
        'status',
        'investors',
    ]

    DEAL_DATA = {
        # Location
        "location-TOTAL_FORMS": 1,
        "location-INITIAL_FORMS": 0,
        "location-MIN_NUM_FORMS": 1,
        "location-MAX_NUM_FORMS": 1,
        "location-0-level_of_accuracy": "Exact location",
        "location-0-location": "Rakhaing-Staat, Myanmar (Birma)",
        "location-0-location-map": "Rakhaing-Staat, Myanmar (Birma)",
        "location-0-point_lat": 19.810093,
        "location-0-point_lon": 93.98784269999999,
        "location-0-target_country": 104,
        # General info
        "id_negotiation_status_0": "Contract signed",
        "id_negotiation_status_1": None,
        "id_negotiation_status_2": None,
        # Contract
        "contract-TOTAL_FORMS": 0,
        "contract-INITIAL_FORMS": 0,
        "contract-MIN_NUM_FORMS": 0,
        "contract-MAX_NUM_FORMS": 0,
        # Data source
        "data_source-TOTAL_FORMS": 1,
        "data_source-INITIAL_FORMS": 0,
        "data_source-MIN_NUM_FORMS": 1,
        "data_source-MAX_NUM_FORMS": 1,
        "data_source-0-type": "Media report",
        # Investor
        "id_operational_stakeholder": 1,
    }

    def setUp(self):
        self.users = {
            'reporter': User.objects.get(username='reporter'),
            'editor': User.objects.get(username='editor'),
            'administrator': User.objects.get(username='administrator'),
        }
        self.groups = {
            'reporter': Group.objects.get(name='Reporters'),
            'editor': Group.objects.get(name='Editors'),
            'administrator': Group.objects.get(name='Administrators'),
        }
        self.factory = RequestFactory()

        # Create group permissions
        # This not possible in fixtures, because permissions and content types are created on run-time
        perm_review_activity = Permission.objects.get(codename='review_activity')
        self.groups['editor'].permissions.add(perm_review_activity)
        perm_add_activity = Permission.objects.get(codename='add_activity')
        perm_change_activity = Permission.objects.get(codename='change_activity')
        perm_delete_activity = Permission.objects.get(codename='delete_activity')
        self.groups['administrator'].permissions.add(perm_review_activity)
        self.groups['administrator'].permissions.add(perm_add_activity)
        self.groups['administrator'].permissions.add(perm_change_activity)
        self.groups['administrator'].permissions.add(perm_delete_activity)

        settings.CELERY_ALWAYS_EAGER = True
        settings.ELASTICSEARCH_INDEX_NAME = 'landmatrix_test'

        es_search.create_index()


class TestAddDeal(BaseTestDeal):

    def is_deal_added(self, activity, user):
        # Check if deal is public
        request = self.factory.get(reverse('deal_detail', kwargs={'deal_id': activity.activity_identifier}))
        request.user = AnonymousUser()
        response = DealDetailView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 200, msg='Deal is not public after approval of Administrator')

        # Check if deal is in latest added log
        request = self.factory.get(reverse('log_added'))
        request.user = user
        response = LogAddedView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Latest Added Log does not work')
        activities = list(response.context_data['activities'])
        self.assertGreaterEqual(len(activities), 1, msg='Wrong list of deals in Latest Added Log of Reporter')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Latest Added Log')

        # Check if deal is in elasticsearch/export
        request = self.factory.get(reverse('export', kwargs={'format': 'xls'}))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = AnonymousUser()
        response = ExportView.as_view()(request, format='xls')
        self.assertEqual(response.status_code, 200, msg='Export does not work')
        wb = load_workbook(BytesIO(response.content), read_only=True)
        ws = wb['Deals']
        activity_identifiers = [row[0].value for row in ws.rows]
        if '#%s' % str(activity.activity_identifier) not in activity_identifiers:
            self.fail('Deal does not appear in export (checked XLS only)')

    def test_add_deal_as_reporter(self):
        # Add deal as reporter
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test add deal",
        })
        request = self.factory.post(reverse('add_deal'), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = AddDealView.as_view()(request)
        self.assertEqual(response.status_code, 302, msg='Add deal does not redirect')

        activity = HistoricalActivity.objects.pending().latest()

        # Check if deal appears in my deals section of reporter
        request = self.factory.get(reverse('manage_for_user'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = ManageForUserView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage My Deals/Investors of Reporter does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage My Deals/Investors of Reporter')
        self.assertEqual(activities[0]['id'], activity.id,
                         msg='Deal does not appear in Manage My Deals/Investors of Reporter')

        # Check if deal NOT appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_adds'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageAddsView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Deals/Investors of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 0, msg='Manage Pending Deals of Administrator should be empty')

        # Check if deal appears in manage section of editor
        request = self.factory.get(reverse('manage_pending_adds'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ManageAddsView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Additions of Editor does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1)
        self.assertEqual(activities[0]['id'], activity.id,
                         msg='Deal does not appear in Manage Pending Deals/Investors of Editor')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username,
                         msg='Deal has wrong user in Manage Pending Deals/Investors of Editor')

        # Approve deal as editor
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Editor",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        self.assertEqual(response.status_code, 302, msg='Approve deal by Editor does not redirect')

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_adds'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageAddsView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Additions of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Additions of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Additions of Administrator')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username, msg='Deal has wrong user in Manage Pending Additions of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        self.assertEqual(response.status_code, 302, msg='Approve deal by Administrator does not redirect')

        self.is_deal_added(activity, self.users['reporter'])

    def test_add_deal_as_editor(self):
        # Add deal as editor
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test add deal",
        })
        request = self.factory.post(reverse('add_deal'), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = AddDealView.as_view()(request)
        self.assertEqual(response.status_code, 302, msg='Add deal does not redirect')

        activity = HistoricalActivity.objects.pending().latest()

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_adds'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageAddsView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Additions of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Additions of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Additions of Administrator')
        self.assertEqual(activities[0]['user'], self.users['editor'].username, msg='Deal has wrong user in Manage Pending Additions of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        self.assertEqual(response.status_code, 302, msg='Approve deal by Administrator does not redirect')

        self.is_deal_added(activity, self.users['editor'])

    def test_add_deal_as_administrator(self):
        # Add deal as administrator
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test add deal",
        })
        request = self.factory.post(reverse('add_deal'), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = AddDealView.as_view()(request)
        self.assertEqual(response.status_code, 302, msg='Add deal does not redirect')

        activity = HistoricalActivity.objects.public().latest()

        self.is_deal_added(activity, self.users['administrator'])


class TestChangeDeal(BaseTestDeal):
    fixtures = [
        'countries_and_regions',
        'users_and_groups',
        'status',
        'activities',
    ]

    def is_deal_changed(self, activity, user):
        # Check if deal is public
        request = self.factory.get(reverse('deal_detail', kwargs={'deal_id': activity.activity_identifier}))
        request.user = AnonymousUser()
        response = DealDetailView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 200, msg='Deal is not public after approval of Administrator')

        # Check if deal is in latest changed log
        request = self.factory.get(reverse('log_modified'))
        request.user = user
        response = LogModifiedView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Latest Modified Log does not work')
        activities = list(response.context_data['activities'])
        self.assertGreaterEqual(len(activities), 1, msg='Wrong list of deals in Latest Modified Log')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Latest Modified Log')

        # Check if deal is in elasticsearch/export
        request = self.factory.get(reverse('export', kwargs={'format': 'xls'}))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = AnonymousUser()
        response = ExportView.as_view()(request, format='xls')
        self.assertEqual(response.status_code, 200, msg='Export does not work')
        wb = load_workbook(BytesIO(response.content), read_only=True)
        ws = wb['Deals']
        activity_identifiers = [row[0].value for row in ws.rows]
        if '#%s' % str(activity.activity_identifier) not in activity_identifiers:
            self.fail('Deal does not appear in export (checked XLS only)')

    def test_change_deal_as_reporter(self):
        # Change deal as reporter
        activity = HistoricalActivity.objects.public().latest()
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test change deal",
        })
        request = self.factory.post(reverse('change_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = ChangeDealView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 302, msg='Change deal does not redirect')

        activity = HistoricalActivity.objects.pending().latest()

        # Check if deal appears in my deals section of reporter
        request = self.factory.get(reverse('manage_for_user'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = ManageForUserView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage My Deals/Investors of Reporter does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage My Deals/Investors of Reporter')
        self.assertEqual(activities[0]['id'], activity.id,
                         msg='Deal does not appear in Manage My Deals/Investors of Reporter')

        # Check if deal NOT appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_updates'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageUpdatesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Updates of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 0, msg='Manage Pending Updates of Administrator should be empty')

        # Check if deal appears in manage section of editor
        request = self.factory.get(reverse('manage_pending_updates'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ManageUpdatesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Updates of Editor does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Updates of Editor')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Updates of Editor')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username, msg='Deals has wrong user in Manage Pending Updates of Editor')

        # Approve deal as editor
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Editor",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        self.assertEqual(response.status_code, 302, msg='Approve deal by Editor does not redirect')

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_updates'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageUpdatesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Updates of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Updates of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Updates of Administrator')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username, msg='Deal has wrong user in Manage Pending Updates of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        self.assertEqual(response.status_code, 302, msg='Approve deal by Administrator does not redirect')

        self.is_deal_changed(activity, self.users['reporter'])

    def test_change_deal_as_editor(self):
        # Change deal as editor
        activity = HistoricalActivity.objects.public().latest()
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test change deal",
        })
        request = self.factory.post(reverse('change_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ChangeDealView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 302, msg='Change deal does not redirect')

        activity = HistoricalActivity.objects.pending().latest()

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_updates'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageUpdatesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Updates of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Updates of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Updates of Administrator')
        self.assertEqual(activities[0]['user'], self.users['editor'].username, msg='Deal has wrong user in Manage Pending Updates of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_change_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityChangeView.as_view()(request, id=activity.id)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Approve deal by Administrator does not redirect')

        self.is_deal_changed(activity, self.users['editor'])

    def test_change_deal_as_administrator(self):
        # Change deal as administrator
        activity = HistoricalActivity.objects.public().latest()
        data = self.DEAL_DATA.copy()
        data.update({
            # Action comment
            "tg_action_comment": "Test change deal",
        })
        request = self.factory.post(reverse('change_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ChangeDealView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 302, msg='Change deal does not redirect')

        activity = HistoricalActivity.objects.public().latest()

        self.is_deal_changed(activity, self.users['administrator'])


class TestDeleteDeal(BaseTestDeal):
    fixtures = [
        'countries_and_regions',
        'users_and_groups',
        'status',
        'activities',
    ]

    def is_deal_deleted(self, activity, user=None):
        # Check if deal is NOT public
        request = self.factory.get(reverse('deal_detail', kwargs={'deal_id': activity.activity_identifier}))
        request.user = AnonymousUser()
        try:
            response = DealDetailView.as_view()(request, deal_id=activity.activity_identifier)
        except Http404:
            pass
        else:
            self.fail("Deal still exists after deletion")

        # Check if deal is in latest deleted log
        request = self.factory.get(reverse('log_deleted'))
        request.user = user
        response = LogDeletedView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Latest Deleted Log does not work')
        activities = list(response.context_data['activities'])
        self.assertGreaterEqual(len(activities), 1, msg='Wrong list of deals in Latest Deleted Log')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Latest Deleted Log')

        # Check if deal is NOT in elasticsearch/export
        request = self.factory.get(reverse('export', kwargs={'format': 'xls'}))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = AnonymousUser()
        response = ExportView.as_view()(request, format='xls')
        self.assertEqual(response.status_code, 200, msg='Export does not work')
        wb = load_workbook(BytesIO(response.content), read_only=True)
        ws = wb['Deals']
        activity_identifiers = [row[0].value for row in ws.rows]
        if '#%s' % str(activity.activity_identifier) in activity_identifiers:
            self.fail('Deal still appears in export after deletion (checked XLS only)')

    def test_delete_deal_as_reporter(self):
        # Delete deal as reporter
        activity = HistoricalActivity.objects.public().latest()
        data = {
            # Action comment
            "tg_action_comment": "Test delete deal",
        }
        request = self.factory.post(reverse('delete_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = DeleteDealView.as_view()(request, deal_id=activity.activity_identifier)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Delete deal by Reporter does not redirect')

        activity = HistoricalActivity.objects.to_delete().latest()

        # Check if deal appears in my deals section of reporter
        request = self.factory.get(reverse('manage_for_user'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['reporter']
        response = ManageForUserView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage My Deals/Investors of Reporter does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage My Deals/Investors of Reporter')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage My Deals/Investors of Reporter')

        # Check if deal NOT appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_deletes'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageDeletesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Deletes of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 0, msg='Manage Pending Deletes of Administrator should be empty')

        # Check if deal appears in manage section of editor
        request = self.factory.get(reverse('manage_pending_deletes'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ManageDeletesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Deletes of Editor does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Deletes of Editor')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Deletes of Editor')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username, msg='Deal has wrong user in Manage Pending Deletes of Editor')

        # Approve deal as editor
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Editor",
        }
        request = self.factory.post(reverse('manage_approve_delete_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = ApproveActivityDeleteView.as_view()(request, id=activity.id)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Approve deal by Editor does not redirect')

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_deletes'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageDeletesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Deletes of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Deletes of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Deletes of Administrator')
        self.assertEqual(activities[0]['user'], self.users['reporter'].username, msg='Deal has wrong user in Manage Pending Deletes of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_delete_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityDeleteView.as_view()(request, id=activity.id)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Approve deal by Administrator does not redirect')

        self.is_deal_deleted(activity, self.users['reporter'])

    def test_delete_deal_as_editor(self):
        # Delete deal as editor
        activity = HistoricalActivity.objects.public().latest()
        data = {
            # Action comment
            "tg_action_comment": "Test delete deal",
        }
        request = self.factory.post(reverse('delete_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = DeleteDealView.as_view()(request, deal_id=activity.activity_identifier)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Delete deal does not redirect')

        activity = HistoricalActivity.objects.to_delete().latest()

        # Check if deal appears in manage section of administrator
        request = self.factory.get(reverse('manage_pending_deletes'))
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ManageDeletesView.as_view()(request)
        self.assertEqual(response.status_code, 200, msg='Manage Pending Deletes of Administrator does not work')
        activities = list(response.context_data['activities'])
        self.assertEqual(len(activities), 1, msg='Wrong list of deals in Manage Pending Deletes of Administrator')
        self.assertEqual(activities[0]['id'], activity.id, msg='Deal does not appear in Manage Pending Deletes of Administrator')
        self.assertEqual(activities[0]['user'], self.users['editor'].username, msg='Deal has wrong user in Manage Pending Deletes of Administrator')

        # Approve deal as administrator
        data = {
            # Action comment
            "tg_action_comment": "Test approve by Administrator",
        }
        request = self.factory.post(reverse('manage_approve_delete_deal', kwargs={'id': activity.id}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = ApproveActivityDeleteView.as_view()(request, id=activity.id)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Approve deal does not redirect')

        self.is_deal_deleted(activity, self.users['editor'])

    def test_delete_deal_as_administrator(self):
        # Delete deal as administrator
        activity = HistoricalActivity.objects.public().latest()
        data = {
            # Action comment
            "tg_action_comment": "Test delete deal",
        }
        request = self.factory.post(reverse('delete_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = DeleteDealView.as_view()(request, deal_id=activity.activity_identifier)
        #if response.status_code == 200:
        #    # For debugging purposes
        #    errors = list(filter(None, [form.errors or None for form in response.context_data['forms']]))
        #    self.assertEqual(errors, [])
        self.assertEqual(response.status_code, 302, msg='Delete deal does not redirect')

        activity = HistoricalActivity.objects.deleted().latest()

        self.is_deal_deleted(activity, self.users['administrator'])

    def test_recover_deal_as_editor(self):
        # Recover deal as editor
        activity = HistoricalActivity.objects.deleted().latest()
        data = {
            # Action comment
            "tg_action_comment": "Test recover deal",
        }
        request = self.factory.post(reverse('recover_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['editor']
        response = RecoverDealView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 302, msg='Recover deal does not redirect')

        # Check if deal is public
        request = self.factory.get(reverse('deal_detail', kwargs={'deal_id': activity.activity_identifier}))
        request.user = AnonymousUser()
        try:
            response = DealDetailView.as_view()(request, deal_id=activity.activity_identifier)
        except Http404:
            pass
        else:
            self.fail("Deal recovered although editors shouldn't be able to recover")

    def test_recover_deal_as_administrator(self):
        # Recover deal as editor
        activity = HistoricalActivity.objects.deleted().latest()
        data = {
            # Action comment
            "tg_action_comment": "Test recover deal",
        }
        request = self.factory.post(reverse('recover_deal', kwargs={'deal_id': activity.activity_identifier}), data)
        # Mock messages framework (not available for unit tests)
        setattr(request, 'session', {})
        messages = FallbackStorage(request)
        setattr(request, '_messages', messages)
        request.user = self.users['administrator']
        response = RecoverDealView.as_view()(request, deal_id=activity.activity_identifier)
        self.assertEqual(response.status_code, 302, msg='Recover deal does not redirect')

        # Check if deal is public
        request = self.factory.get(reverse('deal_detail', kwargs={'deal_id': activity.activity_identifier}))
        request.user = AnonymousUser()
        try:
            response = DealDetailView.as_view()(request, deal_id=activity.activity_identifier)
        except Http404:
            self.fail("Deal still deleted after recovering")
        else:
            pass
