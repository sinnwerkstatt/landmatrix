import pickle
import sys

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

deal_headers = [
    "Is public",
    "Deal scope",
    "Deal size",
    "Current size under contract",
    "Current size in operation (production)",
    "Current negotiation status",
    "Current implementation status",
    # "Fully updated",
    # "Top parent companies",
    "Intended size (in ha)",
    "Size under contract (leased or purchased area, in ha)",
    "Size in operation (production, in ha)",
    "Comment on land area",
    "Intention of investment",
    "Comment on intention of investment",
    "Nature of the deal",
    "Comment on nature of the deal",
    "Negotiation status",
    "Comment on negotiation status",
    "Implementation status",
    "Comment on implementation status",
    "Purchase price",
    "Purchase price currency",
    "Purchase price area type",
    "Purchase price area",
    "Comment on purchase price",
    "Annual leasing fee",
    "Annual leasing fee currency",
    "Annual leasing fee type",
    "Annual leasing fee area",
    "Comment on leasing fees",
    "Contract farming",
    # "On leased / purchased",
    # "On leased / purchased area (in ha)",
    # "On leased / purchased farmers",
    # "On leased / purchased households",
    # "Not on leased / purchased area (out-grower)",
    # "Not on leased / purchased area (out-grower, in ha)",
    # "Not on leased / purchased farmers (out-grower)",
    # "Not on leased / purchased households (out-grower)",
    "Comment on contract farming",
    "Jobs created (total)",
    "Planned number of jobs (total)",
    "Planned employees (total)",
    "Planned daily/seasonal workers (total)",
    # "Current number of jobs (total)",
    # "Current number of employees (total)",
    # "Current number of daily/seasonal workers (total)",
    "Comment on jobs created (total)",
    "Jobs created (foreign)",
    "Planned number of jobs (foreign)",
    "Planned employees (foreign)",
    "Planned daily/seasonal workers (foreign)",
    # "Current number of jobs (foreign)",
    # "Current number of employees (foreign)",
    # "Current number of daily/seasonal workers (foreign)",
    "Comment on jobs created (foreign)",
    "Jobs created (domestic)",
    "Planned number of jobs (domestic)",
    "Planned employees (domestic)",
    "Planned daily/seasonal workers (domestic)",
    # "Current number of jobs (domestic)",
    # "Current number of employees (domestic)",
    # "Current number of daily/seasonal workers (domestic)",
    "Comment on jobs created (domestic)",
    "Actors involved in the negotiation / admission process",
    "Name of investment project",
    "Comment on investment chain",
    "Operating company: Investor ID",
    "Operating company: Name",
    "Operating company: Country of registration/origin",
    "Operating company: Classification",
    "Operating company: Investor homepage",
    "Operating company: Opencorporates link",
    "Operating company: Comment",
    "Name of community",
    "Name of indigenous people",
    "Comment on communities / indigenous peoples affected",
    "Recognition status of community land tenure",
    "Comment on recognitions status of community land tenure",
    "Community consultation",
    "Comment on consultation of local community",
    "Community reaction",
    "Comment on community reaction",
    "Presence of land conflicts",
    "Comment on presence of land conflicts",
    "Displacement of people",
    "Number of people actually displaced",
    "Number of households actually displaced",
    "Number of people displaced out of their community land",
    "Number of people displaced staying on community land",
    'Number of households displaced "only" from their agricultural fields',
    "Number of people facing displacement once project is fully implemented",
    "Comment on displacement of people",
    "Negative impacts for local communities",
    "Comment on negative impacts for local communities",
    "Promised compensation (e.g. for damages or resettlements)",
    "Received compensation (e.g. for damages or resettlements)",
    "Promised benefits for local communities",
    "Comment on promised benefits for local communities",
    "Materialized benefits for local communities",
    "Comment on materialized benefits for local communities",
    "Presence of organizations and actions taken (e.g. farmer organizations, NGOs, etc.)",
    "Former land owner",
    "Comment on former land owner",
    "Former land use",
    "Comment on former land use",
    "Former land cover",
    "Comment on former land cover",
    # "Crops area/yield/export",
    "Comment on crops",
    # "Livestock area/yield/export",
    "Comment on livestock",
    # "Resources area/yield/export",
    "Comment on resources",
    "Contract farming crops",
    "Comment on contract farming crops",
    "Contract farming livestock",
    "Comment on contract farming livestock",
    "Has domestic use",
    "Domestic use",
    "Has export",
    "Export",
    "Country 1",
    "Country 1 ratio",
    "Country 2",
    "Country 2 ratio",
    "Country 3",
    "Country 3 ratio",
    "Comment on use of produce",
    "In country processing of produce",
    "Comment on in country processing of produce",
    "Processing facilities / production infrastructure of the project (e.g. oil mill, ethanol distillery, biomass power plant etc.)",
    "In-country end products of the project",
    "Water extraction envisaged",
    "Comment on water extraction envisaged",
    "Source of water extraction",
    "Comment on source of water extraction",
    "Comment on how much do investors pay for water",
    "Water extraction amount",
    "Comment on how much water is extracted",
    "Use of irrigation infrastructure",
    "Comment on use of irrigation infrastructure",
    "Water footprint of the investment project",
    "Comment on gender-related info",
    "Application of Voluntary Guidelines on the Responsible Governance of Tenure (VGGT)",
    "Comment on VGGT",
    "Application of Principles for Responsible Agricultural Investments (PRAI)",
    "Comment on PRAI",
    "Overall comment",
    # "Not public",
    "Reason",
    "Comment on not public",
]


def ws_to_dict(ws: Worksheet) -> dict:
    old_header = [cell.value for cell in ws[1]]
    old_dict = {}
    ws.calculate_dimension(force=True)
    for x in ws[2 : ws.max_row]:
        values = {}
        for k, v in zip(old_header, x):
            values[k] = v.value
        old_dict[x[0].value] = values
    return old_dict


def cache_workbook(file, pickle_filename):
    try:
        with open(pickle_filename, "rb") as old_pickle:
            print(f"Found cached version of worksheet {file}: {pickle_filename}")
            ws_dict = pickle.load(old_pickle)
    except FileNotFoundError:
        wb_old = load_workbook(filename=file, read_only=True)
        ws_old = wb_old["Deals"]
        print(f"Parsing worksheet {file}..", end="", flush=True)
        ws_dict = ws_to_dict(ws_old)
        with open(pickle_filename, "wb") as old_pickle:
            pickle.dump(ws_dict, old_pickle)
        print(" done.")
    return ws_dict


def parse_hash_shit(hashstring):
    if not hashstring:
        return hashstring
    ret = set()
    for x in hashstring:
        date, curr, rest = x.split("#", maxsplit=2)
        if "#" in rest:
            area, choices = rest.split("#", maxsplit=1)
            if area.endswith(".0"):
                area = area[:-2]
            if "Other (please specify)" in choices:
                choices = choices.replace("Other (please specify)", "Other")
            if " (for wood and fibre)" in choices:
                choices = choices.replace(" (for wood and fibre)", "")
            rest = f"{area}#{choices}"
        if rest.endswith(".0"):
            rest = rest[:-2]
        ret.add(f"{date}##{rest}")
    return ret


def compare(file_old, file_new, field=None):
    print("comparing", file_old, file_new, "for field", field)

    old_dict = cache_workbook(file_old, "xlsx_compare_old.pickle")
    new_dict = cache_workbook(file_new, "xlsx_compare_new.pickle")

    fields = [field] if field else deal_headers

    unknown_deal_numbers = set()

    for field in fields:
        print(">>>", field, "<<<")
        value_mismatch_counter = 0
        for deal_id, content in old_dict.items():
            print(f"\rcomparing {deal_id} ..", end="", flush=True)
            try:
                comp = new_dict[deal_id]
            except KeyError:
                # print(f"could not find {deal_id}")
                unknown_deal_numbers.add(deal_id)
                continue

            if isinstance(content[field], str) and isinstance(comp[field], str):
                content[field] = content[field].strip().replace("\r\n", "\n")
                comp[field] = comp[field].strip().replace("\r\n", "\n")
                if "|" in content[field]:
                    content[field] = set(content[field].split("|"))
                if "|" in comp[field]:
                    comp[field] = set(comp[field].split("|"))
            if isinstance(content[field], (float, int)) or isinstance(
                comp[field], (float, int)
            ):
                try:
                    content[field] = int(float(content[field]))
                except TypeError:
                    pass
                try:
                    comp[field] = int(float(comp[field]))
                except TypeError:
                    pass

            if field in [
                "Size under contract (leased or purchased area, in ha)",
                "Size in operation (production, in ha)",
                "Intention of investment",
                "Negotiation status",
                "Implementation status",
                "Contract farming crops",
                "Contract farming livestock",
            ]:
                if isinstance(content[field], str):
                    content[field] = [content[field]]
                content[field] = parse_hash_shit(content[field])
                if isinstance(comp[field], str):
                    comp[field] = [comp[field]]
                comp[field] = parse_hash_shit(comp[field])

            if content[field] != comp[field]:
                value_mismatch_counter += 1
                print(f"\n{deal_id}: {content[field]} != {comp[field]}")
        if value_mismatch_counter >= 7:  # TODO: watch out, remove this when confident!
            input("\n\nContinue..\n")
        else:
            print("\n\n")
    print()

    print("unknown_deal_numbers", len(unknown_deal_numbers), ":", unknown_deal_numbers)


def compare_special_cases(file_old, file_new):
    old_dict = cache_workbook(file_old, "xlsx_compare_old.pickle")
    new_dict = cache_workbook(file_new, "xlsx_compare_new.pickle")
    print("\n\n\nCOMPARING SPECIAL CASES\n\n\n")

    for deal_id, content in old_dict.items():
        print(f"\rcomparing {deal_id} ..", end="", flush=True)
        try:
            comp = new_dict[deal_id]
        except KeyError:
            # print(f"could not find {deal_id}")
            # unknown_deal_numbers.add(deal_id)
            continue

        for export in ["Crops", "Livestock", "Resources"]:
            coni = (
                content[f"{export} area"],
                content[f"{export} yield"],
                content[f"{export} export"],
            )
            compi = comp[f"{export} area/yield/export"]
            if coni[0] and "|" not in coni[0] and "|" not in compi:
                condate, concurr, conarea, conval = coni[0].split("#")
                comsplit = compi.split("#")
                comdate, comcurr, comarea, comyield, comexport, comval = comsplit

                if condate == comdate and conarea == comarea and conval == comval:
                    continue
            if any(coni) or compi:
                print(coni, "\t\t", compi)


compare(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
compare_special_cases(sys.argv[1], sys.argv[2])
