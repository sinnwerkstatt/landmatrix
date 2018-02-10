try:
    import xml.etree.cElementTree as ET
except ImportError:
    import xml.etree.ElementTree as ET
from xml.dom.minidom import parseString
import unicodecsv as csv
import zipfile
from io import BytesIO
from openpyxl import Workbook
from collections import OrderedDict

from django.http.response import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.views.generic import View

from grid.views.all_deals_view import AllDealsView
from grid.views.table_group_view import TableGroupView
from grid.views.deal_detail_view import DealDetailView
from grid.views.change_deal_view import ChangeDealView
from grid.views.filter_widget_mixin import FilterWidgetMixin
from grid.forms.investor_form import ExportInvestorForm
from grid.forms.parent_investor_formset import InvestorVentureInvolvementForm
from api.views.list_views import ElasticSearchMixin
from landmatrix.models import Activity, InvestorVentureInvolvement
from landmatrix.models.investor import InvestorBase


class ExportView(FilterWidgetMixin, ElasticSearchMixin, View):
    # TODO: XLS is deprecated, should be removed in templates
    FORMATS = ['csv', 'xml', 'xls', 'xlsx']

    def dispatch(self, request, *args, **kwargs):
        data = request.GET.copy()
        if request.GET:
            self.set_country_region_filter(data)
        else:
            self.remove_country_region_filter()
        self.set_default_filters(data)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        format = kwargs.pop('format').lower()
        if format == 'xls':
            format = 'xlsx'

        deal_id = kwargs.pop('deal_id', None)
        if deal_id:
            # Check if activity exists
            activity = Activity.objects.get(activity_identifier=deal_id)
            query = {
                "constant_score" : {
                    "filter" : {
                        "term" : {
                            "activity_identifier": deal_id
                        }
                    }
                }
            }
        else:
            query = self.create_query_from_filters()
        sort = ['activity_identifier',]

        results = {}
        # Search deals
        deals = self.execute_elasticsearch_query(query, doc_type='deal', fallback=False, sort=sort)
        results['deals'] = self.filter_returned_results(deals)

        # Get all involvements
        if deal_id:
            def get_involvements(involvements):
                parents = []
                for involvement in involvements:
                    # Check if there are parent companies for investor
                    parent_involvements = InvestorVentureInvolvement.objects.filter(
                        fk_venture=involvement.fk_investor,
                        fk_venture__fk_status__in=(InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN),
                        fk_investor__fk_status__in=(InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN)
                    )
                    if parent_involvements:
                        parents.extend(get_involvements(parent_involvements))
                    if involvement.fk_investor.fk_status_id in (InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN):
                        parents.append(involvement.id)
                return parents
            query = {
                "ids": {
                    "type": "involvement",
                    "values": get_involvements(activity.investoractivityinvolvement_set.all())
                }
            }
        else:
            query = {}
        sort = ['fk_venture', 'fk_investor']
        results['involvements'] = self.execute_elasticsearch_query(query, doc_type='involvement', fallback=False, sort=sort)

        # Get all investors
        if deal_id:
            def get_investors(investors):
                parents = []
                for investor in investors:
                    # Check if there are parent companies for investor
                    parent_investors = [i.fk_investor for i in InvestorVentureInvolvement.objects.filter(
                        fk_venture=investor,
                        fk_venture__fk_status__in=(InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN),
                        fk_investor__fk_status__in=(InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN)
                    )]
                    if parent_investors:
                        parents.extend(get_investors(parent_investors))
                    if investor.fk_status_id in (InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN):
                        parents.append(investor.id)
                return parents
            query = {
                "ids": {
                    "type": "investor",
                    "values": get_investors([i.fk_investor for i in activity.investoractivityinvolvement_set.all()])
                }
            }
        else:
            query = {}
        sort = ['investor_identifier',]
        results['investors'] = self.execute_elasticsearch_query(query, doc_type='investor', fallback=False, sort=sort)

        if format not in self.FORMATS:
            raise RuntimeError('Download format not recognized: ' + format)

        filename = 'export'
        return getattr(self, 'export_%s' % format)(
            self.get_data(results, format=format),
            "%s.%s" % (filename, format)
        )

    def export_xlsx(self, data, filename):
        response = HttpResponse(content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        wb = Workbook()

        # Deals tab
        ws_deals = wb.get_sheet_by_name('Sheet')
        ws_deals.title = 'Deals'
        ws_deals.append(data['deals']['headers'])
        for item in data['deals']['items']:
            ws_deals.append(item)

        # Involvements tab
        ws_involvements = wb.create_sheet(title='Involvements')
        ws_involvements.append(data['involvements']['headers'])
        for item in data['involvements']['items']:
            ws_involvements.append(item)

        # Investors tab
        ws_investors = wb.create_sheet(title='Investors')
        ws_investors.append(data['investors']['headers'])
        for item in data['investors']['items']:
            ws_investors.append(item)

        wb.save(response)
        return response

    def export_xml(self, data, filename):
        root = ET.Element('data')

        # Deals
        deals = ET.SubElement(root, 'deals')
        for item in data['deals']['items']:
            row = ET.SubElement(deals, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['deals']['headers'][i])

        # Involvements
        involvements = ET.SubElement(root, 'involvements')
        for item in data['involvements']['items']:
            row = ET.SubElement(involvements, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['involvements']['headers'][i])

        # Investors
        investors = ET.SubElement(root, 'investors')
        for item in data['investors']['items']:
            row = ET.SubElement(investors, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['investors']['headers'][i])

        xml = parseString(ET.tostring(root)).toprettyxml()
        response = HttpResponse(xml, content_type='text/xml')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def export_csv(self, data, filename):
        result = BytesIO()
        zip_file = zipfile.ZipFile(result, "w")

        # Deals CSV
        deals_file = BytesIO()
        writer = csv.writer(deals_file, delimiter=";", encoding='utf-8') #encoding='cp1252'
        writer.writerow(data['deals']['headers'])
        for item in data['deals']['items']:
            writer.writerow(item)
        deals_file.seek(0)
        zip_file.writestr('deals.csv', deals_file.getvalue())

        # Involvements CSV
        involvements_file = BytesIO()
        writer = csv.writer(involvements_file, delimiter=";", encoding='utf-8') #encoding='cp1252'
        writer.writerow(data['involvements']['headers'])
        for item in data['involvements']['items']:
            writer.writerow(item)
        involvements_file.seek(0)
        zip_file.writestr('involvements.csv', involvements_file.getvalue())

        # Investors CSV
        investors_file = BytesIO()
        writer = csv.writer(investors_file, delimiter=";", encoding='utf-8')  # encoding='cp1252'
        writer.writerow(data['investors']['headers'])
        for item in data['investors']['items']:
            writer.writerow(item)
        investors_file.seek(0)
        zip_file.writestr('investors.csv', investors_file.getvalue())

        zip_file.close()
        response = HttpResponse(result.getvalue(), content_type='application/x-zip-compressed')
        filename = filename.replace('.csv', '.zip')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def get_data(self, results, format=None):
        """
        Get headers and format the data of the items to a proper download format.
        Returns an array of arrays, each row is an an array of data
        """
        data = {
            'deals': {
                'headers': [],
                'items': [],
                'max': {},
            },
            'involvements': {
                'headers': [],
                'items': [],
            },
            'investors': {
                'headers': [],
                'items': [],
            },
        }
        # Get deal headers and max formset counts
        exclude = []

        headers = [
            str(_('Deal ID')),
            str(_('Is public')),
            str(_('Deal scope')),
            str(_('Deal size')),
            str(_('Current negotiation status')),
            str(_('Fully updated')),
            str(_('Top parent companies')),
        ]
        for form in ChangeDealView.FORMS:
            formset_name = hasattr(form, "form") and form.Meta.name or None
            form = formset_name and form.form or form
            exclude = []
            if hasattr(form, 'exclude_in_export'):
                exclude = form.exclude_in_export
            # Is formset?
            if formset_name:
                # Get item with maximum forms
                form_count = [i.get('%s_count' % formset_name, 0) for i in results['deals']]
                data['deals']['max'][formset_name] = form_count and max(form_count) or 0
                for i in range(0, data['deals']['max'][formset_name]):
                    for field_name, field in form.base_fields.items():
                        if field_name in exclude:
                            continue
                        if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                            continue
                        headers.append('%s %i: %s' % (
                            form.form_title,
                            i + 1,
                            str(field.label),
                        ))
            else:
                for field_name, field in form.base_fields.items():
                    if field_name in exclude:
                        continue
                    if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                        continue
                    headers.append(str(field.label))

            exclude = []
            if hasattr(ExportInvestorForm, 'exclude_in_export'):
                exclude = ExportInvestorForm.exclude_in_export
            # Append operating company attributes to investor info
            if form.Meta.name == 'investor_info':
                for field_name, field in ExportInvestorForm.base_fields.items():
                    if field_name in exclude:
                        continue
                    headers.append('%s: %s' % (_('Operating company'), str(field.label)))
        data['deals']['headers'] = headers

        # Get deals
        rows = []
        for item in results['deals']:
            row = [
                item.get('activity_identifier'),            # ID
                item.get('is_public_display'),              # Is public
                item.get('deal_scope'),                     # Deal Scope
                item.get('deal_size'),                      # Deal Size
                item.get('current_negotiation_status_display'),  # Current negotiation status
                item.get('fully_updated_date'),             # Fully updated date
                item.get('top_investors'),                  # Top investors
            ]
            for form in ChangeDealView.FORMS:
                formset_name = hasattr(form, "form") and form.Meta.name or None
                form = formset_name and form.form or form
                exclude = []
                if hasattr(form, 'exclude_in_export'):
                    exclude = form.exclude_in_export
                # Is formset?
                if formset_name:
                    for i in range(0, data['deals']['max'][formset_name]):
                        for field_name, field in form.base_fields.items():
                            if field_name in exclude:
                                continue
                            if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                                continue
                            row.append(self.get_export_value(field_name, item, formset_index=i, format=format))
                else:
                    for field_name, field in form.base_fields.items():
                        if field_name in exclude:
                            continue
                        if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                            continue
                        row.append(self.get_export_value(field_name, item, format=format))
                # Append operating company attributes to investor info
                if form.Meta.name == 'investor_info':
                    exclude = []
                    if hasattr(ExportInvestorForm, 'exclude_in_export'):
                        exclude = ExportInvestorForm.exclude_in_export
                    for field_name, field in ExportInvestorForm.base_fields.items():
                        if field_name in exclude:
                            continue
                        row.append(self.get_export_value('operating_company_%s' % field_name, item, format=format))
            rows.append(row)
        data['deals']['items'] = rows

        # Get involvement headers
        exclude = []
        if hasattr(InvestorVentureInvolvementForm, 'exclude_in_export'):
            exclude = InvestorVentureInvolvementForm.exclude_in_export
        headers = []
        for field_name, field in InvestorVentureInvolvementForm.base_fields.items():
            if field_name in exclude:
                continue
            headers.append(str(field.label))
        data['involvements']['headers'] = headers
        # Get involvements
        rows = []
        for item in results['involvements']:
            item = item.get('_source', {})
            row = []
            for field_name, field in InvestorVentureInvolvementForm.base_fields.items():
                if field_name in exclude:
                    continue
                row.append(self.get_export_value(field_name, item, format=format))
            rows.append(row)
        data['involvements']['items'] = rows

        # Get investor headers
        exclude = []
        if hasattr(ExportInvestorForm, 'exclude_in_export'):
            exclude = ExportInvestorForm.exclude_in_export
        headers = []
        for field_name, field in ExportInvestorForm.base_fields.items():
            if field_name in exclude:
                continue
            headers.append(str(field.label))
        data['investors']['headers'] = headers
        # Get investors
        rows = []
        for item in results['investors']:
            item = item.get('_source', {})
            row = []
            for field_name, field in ExportInvestorForm.base_fields.items():
                if field_name in exclude:
                    continue
                row.append(self.get_export_value(field_name, item, format=format))
            rows.append(row)
        data['investors']['items'] = rows

        return data

    def get_export_value(self, name, data, formset_index=None, format=None):
        if '%s_display' % name in data:
            value = data.get('%s_display' % name)
        else:
            value = data.get('%s' % name) or ''
        if isinstance(value, (list, tuple)):
            if formset_index is not None:
                try:
                    value = value[formset_index]
                except IndexError:
                    value = ''
            else:
                value = value[0] if len(value) > 0 else ''
        if format == 'xlsx':
            return value.encode('unicode_escape').decode('utf-8')
        else:
            return value

class AllDealsExportView(AllDealsView, ExportView):
    def dispatch(self, request, *args, **kwargs):
        format = kwargs.pop('format')
        kwargs['group'] = 'all'
        context = super().get_context_data(*args, **kwargs)
        return self.export(
            context['data']['items'],
            context['columns'],
            format,
            filename=kwargs['group'])

    def _limit_query(self):
        return False


class TableGroupExportView(TableGroupView, ExportView):
    def dispatch(self, request, *args, **kwargs):
        format = kwargs.pop('format')
        context = super().get_context_data(*args, **kwargs)
        return self.export(
            context['data']['items'],
            context['columns'],
            format,
            filename=kwargs['group'])

    def _limit_query(self):
        return False


class DealDetailExportView(DealDetailView, ExportView):
    def dispatch(self, request, *args, **kwargs):
        format = kwargs.pop('format')
        context = super(DealDetailExportView, self).get_context_data(*args, **kwargs)
        attributes = []
        for form in context['forms']:
            if hasattr(form, 'forms'):
                for i, subform in enumerate(form.forms):
                    for field in subform.get_fields_display():
                        if field['name'].startswith('tg_') and not field['name'].endswith('_comment'):
                            continue
                        label = '%s %i %s' % (
                            subform.form_title,
                            i + 1,
                            field['label']
                        )
                        attributes.append({
                            'field': label,
                            'value': field['value']
                        })
            else:
                for field in form.get_fields_display():
                    if field['name'].startswith('tg_') and not field['name'].endswith('_comment'):
                        continue
                    attributes.append({
                        'field': field['label'],
                        'value': field['value']
                    })
        headers = OrderedDict()
        headers['field'] = {'label': _('Field')}
        headers['value'] = {'label': _('Value')}
        return self.export(attributes, headers, format, filename=kwargs['deal_id'])
