try:
    import xml.etree.cElementTree as ET
except ImportError:  # pragma: no cover
    import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from xml.dom.minidom import parseString

import unicodecsv as csv
from django.http.response import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from apps.api.views.list_views import ElasticSearchMixin
from apps.grid.forms.investor_form import ExportInvestorForm
from apps.grid.forms.parent_investor_formset import InvestorVentureInvolvementForm
from apps.grid.views.deal import DealUpdateView
from apps.grid.views.filter import FilterWidgetMixin
from apps.landmatrix.models import Country, Crop, HistoricalActivity, HistoricalInvestorVentureInvolvement, Region
from apps.landmatrix.models.investor import InvestorBase


# from collections import OrderedDict


class ExportView(FilterWidgetMixin, ElasticSearchMixin, View):
    # TODO: XLS is deprecated, should be removed in templates
    FORMATS = ['csv', 'xml', 'xls', 'xlsx']

    def dispatch(self, request, *args, **kwargs):
        data = request.GET.copy()
        if request.GET:  # pragma: no cover
            self.set_country_region_filter(data)
        else:
            self.remove_country_region_filter()
        self.set_default_filters(data)
        return super().dispatch(request, *args, **kwargs)

    def get_group_value_query(self, query):
        # FIXME: Merge with DealListView.get_group_value_query into Mixin
        group = self.kwargs.get('group', '')
        group_value = self.kwargs.get('group_value', '')
        if not (group and group_value):
            return query

        if not 'bool' in query:  # pragma: no cover
            query['bool'] = {}
        if not 'filter' in query['bool']:  # pragma: no cover
            query['bool']['filter'] = []
        # Deslugify model choices
        if 'country' in group:
            country = Country.objects.get(slug=group_value)
            filter_value = country.id
        elif 'region' in group:
            region = Region.objects.get(slug=group_value)
            filter_value = region.id
        elif 'crop' in group:
            crop = Crop.objects.get(slug=group_value)
            filter_value = crop.id
        else:
            filter_value = group_value
        query['bool']['filter'].append({
            'bool': {
                'filter': {
                    'term': {
                        group: filter_value
                    }
                }
            }
        })
        return query

    def get(self, request, *args, **kwargs):
        format = kwargs.pop('format').lower()
        if format == 'xls':
            format = 'xlsx'

        deal_id = kwargs.pop('deal_id', None)
        if deal_id:
            # Check if activity exists
            queryset = HistoricalActivity.objects
            if request.user.is_authenticated:
                queryset = queryset.public_deleted_or_pending()
            else:
                queryset = queryset.public_or_deleted(self.request.user)
            queryset = queryset.filter(activity_identifier=deal_id).order_by('-id')
            if queryset.count() > 0:
                activity = queryset[0]
            else:  # pragma: no cover
                raise HistoricalActivity.DoesNotExist
            query = {
                "bool": {
                    "must": [
                        {"term" : {"activity_identifier": deal_id}},
                        {"terms": {"status": self.status_list}}
                    ]
                }
            }
        else:
            query = self.create_query_from_filters()
            query = self.get_group_value_query(query)
        sort = ['activity_identifier', ]

        results = {}
        # Search deals
        deals = self.execute_elasticsearch_query(query, doc_type='deal', fallback=False, sort=sort)
        results['deals'] = self.filter_deals(deals)

        # Get all investors
        if deal_id:
            def get_investors(investors):
                parents = []
                for investor in investors:
                    # Check if there are parent companies for investor
                    parent_investors = HistoricalInvestorVentureInvolvement.objects.filter(fk_venture=investor)
                    if not request.user.is_authenticated:
                        parent_investors = parent_investors.filter(
                            fk_venture__fk_status__in=InvestorBase.PUBLIC_STATUSES,
                            fk_investor__fk_status__in=InvestorBase.PUBLIC_STATUSES)
                    parent_investors = [i.fk_investor for i in parent_investors]
                    if parent_investors:
                        parents.extend(get_investors(parent_investors))
                    if request.user.is_authenticated:
                        parents.append(investor.id)
                    elif investor.fk_status_id in InvestorBase.PUBLIC_STATUSES:
                        parents.append(investor.id)
                return parents
            query = {
                "ids": {
                    "type": "investor",
                    "values": get_investors([i.fk_investor for i in activity.involvements.all()])
                }
            }
        else:
            investor_status_list = [InvestorBase.STATUS_ACTIVE, InvestorBase.STATUS_OVERWRITTEN]
            if InvestorBase.STATUS_PENDING in self.status_list:  # pragma: no cover
                investor_status_list += [InvestorBase.STATUS_PENDING]
            if InvestorBase.STATUS_DELETED in self.status_list:  # pragma: no cover
                investor_status_list += [InvestorBase.STATUS_DELETED]
            query = {
                "terms": {"fk_status": investor_status_list}
            }
        sort = ['investor_identifier',]
        investors = self.execute_elasticsearch_query(query, doc_type='investor', fallback=False, sort=sort)
        investors = self.filter_investors(investors)
        results['investors'] = investors

        # Get all involvements
        if deal_id:
            def get_involvements(involvements):
                parents = []
                for involvement in involvements:
                    # Check if there are parent companies for investor
                    parent_involvements = HistoricalInvestorVentureInvolvement.objects.filter(
                        fk_venture=involvement.fk_investor
                    )
                    if not request.user.is_authenticated:
                        parent_involvements = parent_involvements.filter(
                            fk_venture__fk_status__in=InvestorBase.PUBLIC_STATUSES,
                            fk_investor__fk_status__in=InvestorBase.PUBLIC_STATUSES
                        )
                    if parent_involvements:
                        parents.extend(get_involvements(parent_involvements))
                    if request.user.is_authenticated:
                        parents.append(involvement.id)
                    elif involvement.fk_investor.fk_status_id in InvestorBase.PUBLIC_STATUSES:
                        parents.append(involvement.id)
                return parents
            query = {
                "ids": {
                    "type": "involvement",
                    "values": get_involvements(activity.involvements.all())
                }
            }
        else:
            # Get involvements for listed investors
            query = {}
        sort = ['fk_venture', 'fk_investor']
        involvements = self.execute_elasticsearch_query(query, doc_type='involvement', fallback=False, sort=sort)
        results['involvements'] = self.filter_involvements(involvements, investors=investors)

        if format not in self.FORMATS:  # pragma: no cover
            raise RuntimeError('Download format not recognized: ' + format)

        filename = 'export'
        return getattr(self, 'export_%s' % format)(
            self.get_data(results, format=format),
            "%s.%s" % (filename, format)
        )

    def export_xlsx(self, data, filename):
        response = HttpResponse(content_type="application/ms-excel")
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        wb = Workbook()

        # Deals tab
        ws_deals = wb['Sheet']
        ws_deals.title = 'Deals'
        ws_deals.append(data['deals']['headers'])
        for item in data['deals']['items']:
            try:
                ws_deals.append(item)
            except IllegalCharacterError:  # pragma: no cover
                ws_deals.append([str(i).encode('unicode_escape').decode('utf-8') for i in item])

        # Involvements tab
        ws_involvements = wb.create_sheet(title='Involvements')
        ws_involvements.append(data['involvements']['headers'])
        for item in data['involvements']['items']:
            ws_involvements.append(item)

        # Investors tab
        ws_investors = wb.create_sheet(title='Investors')
        ws_investors.append(data['investors']['headers'])
        for item in data['investors']['items']:
            ws_investors.append(item)

        wb.save(response)
        return response

    def export_xml(self, data, filename):
        root = ET.Element('data')

        # Deals
        deals = ET.SubElement(root, 'deals')
        for item in data['deals']['items']:
            row = ET.SubElement(deals, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['deals']['headers'][i])

        # Involvements
        involvements = ET.SubElement(root, 'involvements')
        for item in data['involvements']['items']:
            row = ET.SubElement(involvements, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['involvements']['headers'][i])

        # Investors
        investors = ET.SubElement(root, 'investors')
        for item in data['investors']['items']:
            row = ET.SubElement(investors, "item")
            for i, value in enumerate(item):
                field = ET.SubElement(row, "field")
                field.text = str(value)
                field.set("name", data['investors']['headers'][i])
        xml = ET.tostring(root)
        try:
            xml = parseString(xml).toprettyxml()
        except:  # pragma: no cover
            pass
        response = HttpResponse(xml, content_type='text/xml')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def export_csv(self, data, filename):
        result = BytesIO()
        zip_file = zipfile.ZipFile(result, "w")

        # Deals CSV
        deals_file = BytesIO()
        writer = csv.writer(deals_file, delimiter=";", encoding='utf-8') #encoding='cp1252'
        writer.writerow(data['deals']['headers'])
        for item in data['deals']['items']:
            writer.writerow(item)
        deals_file.seek(0)
        zip_file.writestr('deals.csv', deals_file.getvalue())

        # Involvements CSV
        involvements_file = BytesIO()
        writer = csv.writer(involvements_file, delimiter=";", encoding='utf-8') #encoding='cp1252'
        writer.writerow(data['involvements']['headers'])
        for item in data['involvements']['items']:
            writer.writerow(item)
        involvements_file.seek(0)
        zip_file.writestr('involvements.csv', involvements_file.getvalue())

        # Investors CSV
        investors_file = BytesIO()
        writer = csv.writer(investors_file, delimiter=";", encoding='utf-8')  # encoding='cp1252'
        writer.writerow(data['investors']['headers'])
        for item in data['investors']['items']:
            writer.writerow(item)
        investors_file.seek(0)
        zip_file.writestr('investors.csv', investors_file.getvalue())

        zip_file.close()
        response = HttpResponse(result.getvalue(), content_type='application/x-zip-compressed')
        filename = filename.replace('.csv', '.zip')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def get_data(self, results, format=None):
        """
        Get headers and format the data of the items to a proper download format.
        Returns an array of arrays, each row is an an array of data
        """
        data = {
            'deals': {
                'headers': [],
                'items': [],
                'max': {},
            },
            'involvements': {
                'headers': [],
                'items': [],
            },
            'investors': {
                'headers': [],
                'items': [],
            },
        }
        # Get deal headers and max formset counts
        exclude = []

        headers = [
            str(_('Deal ID')),
            str(_('Is public')),
            str(_('Deal scope')),
            str(_('Deal size')),
            str(_('Current size under contract')),
            str(_('Current size in operation (production)')),
            str(_('Current negotiation status')),
            str(_('Current implementation status')),
            str(_('Fully updated')),
            str(_('Top parent companies')),
        ]
        for form in DealUpdateView.FORMS:
            formset_name = hasattr(form, "form") and form.Meta.name or None
            form = formset_name and form.form or form
            exclude = []
            if hasattr(form, 'exclude_in_export'):
                exclude = form.exclude_in_export
            # Is formset?
            if formset_name:
                # Get item with maximum forms
                form_count = [i.get('%s_count' % formset_name, 0) for i in results['deals']]
                data['deals']['max'][formset_name] = form_count and max(form_count) or 0
                for i in range(0, data['deals']['max'][formset_name]):
                    for field_name, field in form.base_fields.items():
                        if field_name in exclude:
                            continue
                        if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                            continue
                        headers.append('%s %i: %s' % (
                            form.form_title,
                            i + 1,
                            str(field.label),
                        ))
            else:
                for field_name, field in form.base_fields.items():
                    if field_name in exclude:
                        continue
                    if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                        continue
                    headers.append(str(field.label))

            exclude = []
            if hasattr(ExportInvestorForm, 'exclude_in_export'):
                exclude = ExportInvestorForm.exclude_in_export
            # Append operating company attributes to investor info
            if form.Meta.name == 'investor_info':
                for field_name, field in ExportInvestorForm.base_fields.items():
                    if field_name in exclude:
                        continue
                    headers.append('%s: %s' % (_('Operating company'), str(field.label)))
        data['deals']['headers'] = headers

        # Get deals
        rows = []
        for item in results['deals']:
            row = [
                item.get('activity_identifier'),
                item.get('is_public_display'),
                item.get('deal_scope'),
                item.get('deal_size'),
                item.get('current_contract_size'),
                item.get('current_production_size'),
                item.get('current_negotiation_status_display'),
                item.get('current_implementation_status_display'),
                item.get('fully_updated_date'),
                item.get('top_investors'),
            ]
            for form in DealUpdateView.FORMS:
                formset_name = hasattr(form, "form") and form.Meta.name or None
                form = formset_name and form.form or form
                exclude = []
                if hasattr(form, 'exclude_in_export'):
                    exclude = form.exclude_in_export
                # Is formset?
                if formset_name:
                    for i in range(0, data['deals']['max'][formset_name]):
                        for field_name, field in form.base_fields.items():
                            if field_name in exclude:
                                continue
                            if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                                continue
                            row.append(self.get_export_value(field_name, item, formset_index=i, format=format))
                else:
                    for field_name, field in form.base_fields.items():
                        if field_name in exclude:
                            continue
                        if field_name.startswith('tg_') and not field_name.endswith('_comment'):
                            continue
                        row.append(self.get_export_value(field_name, item, format=format))
                # Append operating company attributes to investor info
                if form.Meta.name == 'investor_info':
                    exclude = []
                    if hasattr(ExportInvestorForm, 'exclude_in_export'):
                        exclude = ExportInvestorForm.exclude_in_export
                    for field_name, field in ExportInvestorForm.base_fields.items():
                        if field_name in exclude:
                            continue
                        row.append(self.get_export_value('operating_company_%s' % field_name, item, format=format))
            rows.append(row)
        data['deals']['items'] = rows

        # Get involvement headers
        exclude = []
        if hasattr(InvestorVentureInvolvementForm, 'exclude_in_export'):
            exclude = InvestorVentureInvolvementForm.exclude_in_export
        headers = []
        for field_name, field in InvestorVentureInvolvementForm.base_fields.items():
            if field_name in exclude:
                continue
            headers.append(str(field.label))
            # Add investor name after ID
            if field_name in ('fk_venture', 'fk_investor'):
                label = str(field.label).replace('ID', 'Name')
                headers.append(str(label))
        data['involvements']['headers'] = headers
        # Get involvements
        rows = []
        for item in results['involvements']:
            row = []
            for field_name, field in InvestorVentureInvolvementForm.base_fields.items():
                if field_name in exclude:
                    continue
                row.append(self.get_export_value(field_name, item, format=format))

                # Add investor name after ID
                if field_name in ('fk_venture', 'fk_investor'):
                    field_name += '_name'
                    row.append(self.get_export_value(field_name, item, format=format))
            rows.append(row)
        data['involvements']['items'] = rows

        # Get investor headers
        exclude = []
        if hasattr(ExportInvestorForm, 'exclude_in_export'):
            exclude = ExportInvestorForm.exclude_in_export
        headers = []
        for field_name, field in ExportInvestorForm.base_fields.items():
            if field_name in exclude:
                continue
            headers.append(str(field.label))
        data['investors']['headers'] = headers
        # Get investors
        rows = []
        for item in results['investors']:
            row = []
            for field_name, field in ExportInvestorForm.base_fields.items():
                if field_name in exclude:
                    continue
                row.append(self.get_export_value(field_name, item, format=format))
            rows.append(row)
        data['investors']['items'] = rows

        return data

    def get_export_value(self, name, data, formset_index=None, format=None):
        if '%s_display' % name in data:
            value = data.get('%s_display' % name)
        else:
            value = data.get('%s' % name) or ''
        if isinstance(value, (list, tuple)):
            if formset_index is not None:
                try:
                    value = value[formset_index]
                except IndexError:
                    value = ''
            else:
                value = value[0] if len(value) > 0 else ''
        return value
