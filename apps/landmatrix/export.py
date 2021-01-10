import json
import zipfile
from io import BytesIO

import unicodecsv as csv
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from apps.graphql.tools import parse_filters
from apps.landmatrix.models import (
    Deal,
    Investor,
    InvestorVentureInvolvement,
    Currency,
    Location,
    DataSource,
    Contract,
    Country,
    Crop,
    Animal,
    Mineral,
)
from apps.landmatrix.utils import InvolvementNetwork
from apps.utils import qs_values_to_dict, arrayfield_choices_display

deal_fields = {
    "id": "Deal ID",
    "is_public": "Is public",
    "transnational": "Deal scope",
    "deal_size": "Deal size",
    "current_contract_size": "Current size under contract",
    "current_production_size": "Current size in operation (production)",
    "current_negotiation_status": "Current negotiation status",
    "current_implementation_status": "Current implementation status",
    "fully_updated_at": "Fully updated",
    "top_investors": "Top parent companies",
    "intended_size": "Intended size (in ha)",
    "contract_size": "Size under contract (leased or purchased area, in ha)",
    "production_size": "Size in operation (production, in ha)",
    "land_area_comment": "Comment on land area",
    "intention_of_investment": "Intention of investment",
    "intention_of_investment_comment": "Comment on intention of investment",
    "nature_of_deal": "Nature of the deal",
    "nature_of_deal_comment": "Comment on nature of the deal",
    "negotiation_status": "Negotiation status",
    "negotiation_status_comment": "Comment on negotiation status",
    "implementation_status": "Implementation status",
    "implementation_status_comment": "Comment on implementation status",
    "purchase_price": "Purchase price",
    "purchase_price_currency": "Purchase price currency",
    "purchase_price_type": "Purchase price area type",
    "purchase_price_area": "Purchase price area",
    "purchase_price_comment": "Comment on purchase price",
    "annual_leasing_fee": "Annual leasing fee",
    "annual_leasing_fee_currency": "Annual leasing fee currency",
    "annual_leasing_fee_type": "Annual leasing fee type",
    "annual_leasing_fee_area": "Annual leasing fee area",
    "annual_leasing_fee_comment": "Comment on leasing fees",
    "contract_farming": "Contract farming",
    "on_the_lease": "On leased / purchased area",
    "on_the_lease_area": "On leased / purchased area (in ha)",
    "on_the_lease_farmers": "On leased / purchased farmers",
    "on_the_lease_households": "On leased / purchased households",
    "off_the_lease": "Not on leased / purchased area (out-grower)",
    "off_the_lease_area": "Not on leased / purchased area (out-grower, in ha)",
    "off_the_lease_farmers": "Not on leased / purchased farmers (out-grower)",
    "off_the_lease_households": "Not on leased / purchased households (out-grower)",
    "contract_farming_comment": "Comment on contract farming",
    "total_jobs_created": "Jobs created (total)",
    "total_jobs_planned": "Planned number of jobs (total)",
    "total_jobs_planned_employees": "Planned employees (total)",
    "total_jobs_planned_daily_workers": "Planned daily/seasonal workers (total)",
    "total_jobs_current": "Current number of jobs (total)",
    "total_jobs_current_employees": "Current number of employees (total)",
    "total_jobs_current_daily_workers": "Current number of daily/seasonal workers (total)",
    "total_jobs_created_comment": "Comment on jobs created (total)",
    "foreign_jobs_created": "Jobs created (foreign)",
    "foreign_jobs_planned": "Planned number of jobs (foreign)",
    "foreign_jobs_planned_employees": "Planned employees (foreign)",
    "foreign_jobs_planned_daily_workers": "Planned daily/seasonal workers (foreign)",
    "foreign_jobs_current": "Current number of jobs (foreign)",
    "foreign_jobs_current_employees": "Current number of employees (foreign)",
    "foreign_jobs_current_daily_workers": "Current number of daily/seasonal workers (foreign)",
    "foreign_jobs_created_comment": "Comment on jobs created (foreign)",
    "domestic_jobs_created": "Jobs created (domestic)",
    "domestic_jobs_planned": "Planned number of jobs (domestic)",
    "domestic_jobs_planned_employees": "Planned employees (domestic)",
    "domestic_jobs_planned_daily_workers": "Planned daily/seasonal workers (domestic)",
    "domestic_jobs_current": "Current number of jobs (domestic)",
    "domestic_jobs_current_employees": "Current number of employees (domestic)",
    "domestic_jobs_current_daily_workers": "Current number of daily/seasonal workers (domestic)",
    "domestic_jobs_created_comment": "Comment on jobs created (domestic)",
    "involved_actors": "Actors involved in the negotiation / admission process",
    "project_name": "Name of investment project",
    "investment_chain_comment": "Comment on investment chain",
    "operating_company__id": "Operating company: Investor ID",
    "operating_company__name": "Operating company: Name",
    "operating_company__country__name": "Operating company: Country of registration/origin",
    "operating_company__classification": "Operating company: Classification",
    "operating_company__homepage": "Operating company: Investor homepage",
    "operating_company__opencorporates": "Operating company: Opencorporates link",
    "operating_company__comment": "Operating company: Comment",
    "name_of_community": "Name of community",
    "name_of_indigenous_people": "Name of indigenous people",
    "people_affected_comment": "Comment on communities / indigenous peoples affected",
    "recognition_status": "Recognition status of community land tenure",
    "recognition_status_comment": "Comment on recognitions status of community land tenure",
    "community_consultation": "Community consultation",
    "community_consultation_comment": "Comment on consultation of local community",
    "community_reaction": "Community reaction",
    "community_reaction_comment": "Comment on community reaction",
    "land_conflicts": "Presence of land conflicts",
    "land_conflicts_comment": "Comment on presence of land conflicts",
    "displacement_of_people": "Displacement of people",
    "displaced_people": "Number of people actually displaced",
    "displaced_households": "Number of households actually displaced",
    "displaced_people_from_community_land": "Number of people displaced out of their community land",
    "displaced_people_within_community_land": "Number of people displaced staying on community land",
    "displaced_households_from_fields": 'Number of households displaced "only" from their agricultural fields',
    "displaced_people_on_completion": "Number of people facing displacement once project is fully implemented",
    "displacement_of_people_comment": "Comment on displacement of people",
    "negative_impacts": "Negative impacts for local communities",
    "negative_impacts_comment": "Comment on negative impacts for local communities",
    "promised_compensation": "Promised compensation (e.g. for damages or resettlements)",
    "received_compensation": "Received compensation (e.g. for damages or resettlements)",
    "promised_benefits": "Promised benefits for local communities",
    "promised_benefits_comment": "Comment on promised benefits for local communities",
    "materialized_benefits": "Materialized benefits for local communities",
    "materialized_benefits_comment": "Comment on materialized benefits for local communities",
    "presence_of_organizations": "Presence of organizations and actions taken (e.g. farmer organizations, NGOs, etc.)",
    "former_land_owner": "Former land owner",
    "former_land_owner_comment": "Comment on former land owner",
    "former_land_use": "Former land use",
    "former_land_use_comment": "Comment on former land use",
    "former_land_cover": "Former land cover",
    "former_land_cover_comment": "Comment on former land cover",
    "crops": "Crops area/yield/export",
    "crops_comment": "Comment on crops",
    "animals": "Livestock area/yield/export",
    "animals_comment": "Comment on livestock",
    "resources": "Resources area/yield/export",
    "resources_comment": "Comment on resources",
    "contract_farming_crops": "Contract farming crops",
    "contract_farming_crops_comment": "Comment on contract farming crops",
    "contract_farming_animals": "Contract farming livestock",
    "contract_farming_animals_comment": "Comment on contract farming livestock",
    "has_domestic_use": "Has domestic use",
    "domestic_use": "Domestic use",
    "has_export": "Has export",
    "export": "Export",
    "export_country1": "Country 1",
    "export_country1_ratio": "Country 1 ratio",
    "export_country2": "Country 2",
    "export_country2_ratio": "Country 2 ratio",
    "export_country3": "Country 3",
    "export_country3_ratio": "Country 3 ratio",
    "use_of_produce_comment": "Comment on use of produce",
    "in_country_processing": "In country processing of produce",
    "in_country_processing_comment": "Comment on in country processing of produce",
    "in_country_processing_facilities": "Processing facilities / production infrastructure of the project (e.g. oil mill, ethanol distillery, biomass power plant etc.)",
    "in_country_end_products": "In-country end products of the project",
    "water_extraction_envisaged": "Water extraction envisaged",
    "water_extraction_envisaged_comment": "Comment on water extraction envisaged",
    # "source_of_water_extraction": "Source of water extraction",
    "source_of_water_extraction_comment": "Comment on source of water extraction",
    "how_much_do_investors_pay_comment": "Comment on how much do investors pay for water",
    # "water_extraction_amount": "Water extraction amount",
    "water_extraction_amount_comment": "Comment on how much water is extracted",
    # "use_of_irrigation_infrastructure": "Use of irrigation infrastructure",
    "use_of_irrigation_infrastructure_comment": "Comment on use of irrigation infrastructure",
    # "water_footprint": "Water footprint of the investment project",
    "gender_related_information": "Comment on gender-related info",
    # "vggt_applied": "Application of Voluntary Guidelines on the Responsible Governance of Tenure (VGGT)",
    "vggt_applied_comment": "Comment on VGGT",
    # "prai_applied": "Application of Principles for Responsible Agricultural Investments (PRAI)",
    "prai_applied_comment": "Comment on PRAI",
    "overall_comment": "Overall comment",
    "confidential": "Not public",
    # "confidential_reason": "Reason",
    "confidential_comment": "Comment on not public",
}

location_fields = {
    "deal_id": "Deal ID",
    "level_of_accuracy": "Spatial accuracy level",
    "name": "Location",
    "point": "Point",
    "facility_name": "Facility name",
    "description": "Location description",
    "comment": "Comment on location",
}
contract_fields = {
    "deal_id": "Deal ID",
    "number": "Contract number",
    "date": "Contract date",
    "expiration_date": "Contract expiration date",
    "agreement_duration": "Duration of the agreement (in years)",
    "comment": "Comment on contract",
}

datasource_fields = {
    "deal_id": "Deal ID",
    "type": "Data source type",
    "url": "URL",
    "file": "File",
    "publication_title": "Publication title",
    "date": "Date",
    "name": "Name",
    "company": "Organisation",
    "email": "Email",
    "phone": "Phone",
    "open_land_contracts_id": "Open Contracting ID",
    "comment": "Comment on data source",
}

current_negotiation_status_map = {
    "EXPRESSION_OF_INTEREST": "Intended (Expression of interest)",
    "UNDER_NEGOTIATION": "Intended (Under negotiation)",
    "MEMORANDUM_OF_UNDERSTANDING": "Intended (Memorandum of understanding)",
    "ORAL_AGREEMENT": "Concluded (Oral Agreement)",
    "CONTRACT_SIGNED": "Concluded (Contract signed)",
    "NEGOTIATIONS_FAILED": "Failed (Negotiations failed)",
    "CONTRACT_CANCELED": "Failed (Contract cancelled)",
    "CONTRACT_EXPIRED": "Contract expired",
    "CHANGE_OF_OWNERSHIP": "Change of ownership",
    None: "None",
}

current_implementation_status_map = {
    "PROJECT_NOT_STARTED": "Project not started",
    "STARTUP_PHASE": "Startup phase (no production)",
    "IN_OPERATION": "In operation (production)",
    "PROJECT_ABANDONED": "Project abandoned",
    None: "None",
}
intention_of_investment_map = {
    "BIOFUELS": "Biofuels",
    "FOOD_CROPS": "Food crops",
    "FODDER": "Fodder",
    "LIVESTOCK": "Livestock",
    "NON_FOOD_AGRICULTURE": "Non-food agricultural commodities",
    "AGRICULTURE_UNSPECIFIED": "Agriculture unspecified",
    "TIMBER_PLANTATION": "Timber plantation",
    "FOREST_LOGGING": "Forest logging / management",
    "CARBON": "For carbon sequestration/REDD",
    "FORESTRY_UNSPECIFIED": "Forestry unspecified",
    "MINING": "Mining",
    "OIL_GAS_EXTRACTION": "Oil / Gas extraction",
    "TOURISM": "Tourism",
    "INDUSTRY": "Industry",
    "CONVERSATION": "Conservation",
    "LAND_SPECULATION": "Land speculation",
    "RENEWABLE_ENERGY": "Renewable Energy",
    "OTHER": "Other",
}

deal_choices_fields = {
    "intention_of_investment": intention_of_investment_map,
    "investor_classification": dict(Investor.CLASSIFICATION_CHOICES),
}


class Choices:
    choices = {}

    def get(self, name):
        if not self.choices.get(name):
            if name == "currency":
                self.choices[name] = dict(Currency.objects.values_list("id", "name"))
            if name == "country":
                self.choices[name] = dict(Country.objects.values_list("id", "name"))
            if name == "crops":
                self.choices[name] = dict(Crop.objects.values_list("code", "name"))
            if name == "animals":
                self.choices[name] = dict(Animal.objects.values_list("code", "name"))
            if name == "resources":
                self.choices[name] = dict(Mineral.objects.values_list("code", "name"))
        return self.choices[name]


mchoices = Choices()

deal_sub_fields = {
    "top_investors": [
        "top_investors__id",
        "top_investors__name",
        "top_investors__country__name",
    ]
}
deal_flattened_fields = []
for f in deal_fields.keys():
    if f in deal_sub_fields:
        for sf in deal_sub_fields[f]:
            deal_flattened_fields.append(sf)
    else:
        deal_flattened_fields.append(f)

investor_headers = [
    "Investor ID",
    "Name",
    "Country of registration/origin",
    "Classification",
    "Investor homepage",
    "Opencorporates link",
    "Comment",
    "Action comment",
]
investor_fields = [
    "id",
    "name",
    "country__name",
    "classification",
    "homepage",
    "opencorporates",
    "comment",
]
investor_choices_fields = {"classification": dict(Investor.CLASSIFICATION_CHOICES)}

involvement_headers = [
    "Involvement ID",
    "Investor ID Downstream",
    "Investor Name Downstream",
    "Investor ID Upstream",
    "Investor Name Upstream",
    "Relation type",
    "Investment type",
    "Ownership share",
    "Loan amount",
    "Loan currency",
    "Loan date",
    "Comment",
]
involvement_fields = [
    "id",
    "venture_id",
    "venture__name",
    "investor_id",
    "investor__name",
    "role",
    "investment_type",
    "percentage",
    "loans_amount",
    "loans_currency",
    "loans_date",
    "comment",
]
involvement_choices_fields = {"role": dict(InvestorVentureInvolvement.ROLE_CHOICES)}


def flatten_date_current_value(data, field) -> None:
    if not data.get(field):
        return
    data[field] = "|".join(
        [
            "#".join(
                [
                    str(x.get("date", "")),
                    "current" if x.get("current") else "",
                    x["value"] if isinstance(x["value"], str) else f"{x['value']:.0f}",
                ]
            )
            for x in data[field]
            if x.get("value") is not None
        ]
    )


def flatten_array_choices(data, field, choices) -> None:
    if not data.get(field):
        return

    data[field] = "|".join([choices[x] for x in data[field]])


def single_choice(data, field, choices) -> None:
    if not data.get(field):
        return
    data[field] = choices[data[field]]


def bool_cast(data, field) -> None:
    if data.get(field) is None:
        return
    data[field] = "Yes" if data[field] else "No"


class DataDownload:
    def __init__(self, request):
        self.request = request
        self.user = request.user
        deal_id = self.request.GET.get("deal_id")
        filters = self.request.GET.get("filters")
        self.return_format = self.request.GET.get("format", "html")

        if deal_id:
            self._single_deal(deal_id)
        else:
            self._multiple_deals(filters)

    def _single_deal(self, deal_id):
        qs = Deal.objects.filter(id=deal_id)
        deal = qs[0]
        self.deals = [
            self.deal_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                qs, deal_flattened_fields, deal_sub_fields.keys()
            )
        ]
        self.locations = [
            self.location_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                deal.locations.all(), list(location_fields.keys())
            )
        ]
        self.contracts = [
            self.contracts_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                deal.contracts.all(), list(contract_fields.keys())
            )
        ]
        self.datasources = [
            self.datasource_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                deal.datasources.all(), list(datasource_fields.keys())
            )
        ]

        self.investors = []
        self.involvements = []
        if deal.operating_company:
            (
                self.investors,
                self.involvements,
            ) = InvolvementNetwork().flat_view_for_download(deal.operating_company)
        self.filename = f"deal_{deal_id}"

    def _multiple_deals(self, filters):

        qs = Deal.objects.visible(self.user, subset="ACTIVE").order_by("id")
        if filters:
            qs = qs.filter(parse_filters(json.loads(filters)))

        deal_ids = qs.values_list("id", flat=True)

        self.deals = [
            self.deal_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                qs, deal_flattened_fields, deal_sub_fields.keys()
            )
        ]

        self.locations = [
            self.location_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                Location.objects.filter(deal_id__in=deal_ids).order_by("deal_id", "id"),
                list(location_fields.keys()),
            )
        ]
        self.contracts = [
            self.contracts_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                Contract.objects.filter(deal_id__in=deal_ids).order_by("deal_id", "id"),
                list(contract_fields.keys()),
            )
        ]

        self.datasources = [
            self.datasource_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(
                DataSource.objects.filter(deal_id__in=deal_ids).order_by(
                    "deal_id", "id"
                ),
                list(datasource_fields.keys()),
            )
        ]

        qs = Investor.objects.visible(self.user).order_by("id")
        self.investors = [
            self.investor_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(qs, investor_fields, [])
        ]

        qs = InvestorVentureInvolvement.objects.visible(self.user).order_by("id")
        self.involvements = [
            self.involvement_download_format(qs_dict)
            for qs_dict in qs_values_to_dict(qs, involvement_fields, [])
        ]
        self.filename = "export"

    def get_response(self):
        if self.return_format == "xlsx":
            return self.xlsx()
        if self.return_format == "csv":
            return self.csv()
        return self.html()

    def html(self):
        ctx = {
            "deal_headers": deal_fields.values(),
            "deals": self.deals,
            "location_headers": location_fields.values(),
            "locations": self.locations,
            "contract_headers": contract_fields.values(),
            "contracts": self.contracts,
            "datasource_headers": datasource_fields.values(),
            "datasources": self.datasources,
            "investor_headers": investor_headers,
            "investors": self.investors,
            "involvement_headers": involvement_headers,
            "involvements": self.involvements,
        }
        return render(self.request, "landmatrix/export_table.html", ctx)

    def xlsx(self):
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.xlsx"'
        wb = Workbook(write_only=True)

        # Deals tab
        ws_deals = wb.create_sheet(title="Deals")
        ws_deals.append(list(deal_fields.values()))
        for item in self.deals:
            try:
                ws_deals.append(item)
            except IllegalCharacterError:  # pragma: no cover
                ws_deals.append(
                    [str(i).encode("unicode_escape").decode("utf-8") for i in item]
                )

        ## Locations tab
        ws = wb.create_sheet(title="Locations")
        ws.append(list(location_fields.values()))
        [ws.append(item) for item in self.locations]

        ## Contracts tab
        ws = wb.create_sheet(title="Contracts")
        ws.append(list(contract_fields.values()))
        [ws.append(item) for item in self.contracts]

        ## DataSources tab
        ws = wb.create_sheet(title="Data Sources")
        ws.append(list(datasource_fields.values()))
        [ws.append(item) for item in self.datasources]

        # Involvements tab
        ws_involvements = wb.create_sheet(title="Involvements")
        ws_involvements.append(involvement_headers)
        for item in self.involvements:
            ws_involvements.append(item)

        # Investors tab
        ws_investors = wb.create_sheet(title="Investors")
        ws_investors.append(investor_headers)
        for item in self.investors:
            ws_investors.append(item)

        wb.save(response)
        return response

    @staticmethod
    def _csv_writer(data):
        # Deals CSV
        file = BytesIO()
        writer = csv.writer(file, delimiter=";")  # encoding='cp1252'
        for item in data:
            writer.writerow(item)
        file.seek(0)
        return file.getvalue()

    def csv(self):
        result = BytesIO()
        zip_file = zipfile.ZipFile(result, "w")

        # Deals CSV
        deals_data = [deal_fields.values()] + self.deals
        zip_file.writestr("deals.csv", self._csv_writer(deals_data))

        # Locations CSV
        zip_file.writestr(
            "locations.csv",
            self._csv_writer([location_fields.values()] + self.locations),
        )
        # Contracts CSV
        zip_file.writestr(
            "contracts.csv",
            self._csv_writer([contract_fields.values()] + self.contracts),
        )
        # Datasources CSV
        zip_file.writestr(
            "datasources.csv",
            self._csv_writer([datasource_fields.values()] + self.datasources),
        )

        # Involvements CSV
        involvements_data = [involvement_headers] + self.involvements
        zip_file.writestr("involvements.csv", self._csv_writer(involvements_data))

        # Investors CSV
        investors_data = [investor_headers] + self.investors
        zip_file.writestr("investors.csv", self._csv_writer(investors_data))

        zip_file.close()
        response = HttpResponse(
            result.getvalue(), content_type="application/x-zip-compressed"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.filename}.zip"'
        return response

    @staticmethod
    def deal_download_format(data):
        bool_cast(data, "is_public")
        if "transnational" in data:
            data["transnational"] = (
                "transnational" if data["transnational"] else "domestic"
            )

        # flatten top investors
        data["top_investors"] = "|".join(
            [
                "#".join(
                    [
                        ti["name"].replace("#", "").replace("\n", "").strip(),
                        str(ti["id"]),
                        ti["country__name"] if "country__name" in ti else "",
                    ]
                )
                for ti in sorted(data["top_investors"], key=lambda x: x["id"])
            ]
        )
        # map operating company fields
        if "operating_company" in data:
            data["operating_company__id"] = data["operating_company"]["id"]
            data["operating_company__name"] = data["operating_company"]["name"]
            if "country" in data["operating_company"]:
                data["operating_company__country__name"] = data["operating_company"][
                    "country"
                ]["name"]
            if "classification" in data["operating_company"]:
                data["operating_company__classification"] = str(
                    deal_choices_fields["investor_classification"][
                        data["operating_company"]["classification"]
                    ]
                )
            data["operating_company__homepage"] = data["operating_company"]["homepage"]
            data["operating_company__opencorporates"] = data["operating_company"][
                "opencorporates"
            ]
            data["operating_company__comment"] = data["operating_company"]["comment"]

        data["deal_size"] = int(data.get("deal_size", 0))
        data["current_contract_size"] = int(data.get("current_contract_size", 0))
        data["current_production_size"] = int(data.get("current_production_size", 0))

        imp_stat = data.get("current_implementation_status")
        data["current_implementation_status"] = current_implementation_status_map.get(
            imp_stat, imp_stat
        )

        neg_stat = data.get("current_negotiation_status")
        data["current_negotiation_status"] = current_negotiation_status_map.get(
            neg_stat, neg_stat
        )

        fully_updated_at = data.get("fully_updated_at")
        if fully_updated_at:
            data["fully_updated_at"] = fully_updated_at.isoformat()

        flatten_date_current_value(data, "contract_size")
        flatten_date_current_value(data, "production_size")

        if data.get("intention_of_investment"):
            data["intention_of_investment"] = "|".join(
                [
                    "#".join(
                        [
                            str(x.get("date", "")),
                            "current" if x.get("current") else "",
                            x.get("size", ""),
                            ", ".join(
                                deal_choices_fields["intention_of_investment"][y]
                                for y in x.get("value", [])
                            ),
                        ]
                    )
                    for x in data["intention_of_investment"]
                    if x.get("value") is not None
                ]
            )

        flatten_array_choices(data, "nature_of_deal", dict(Deal.NATURE_OF_DEAL_CHOICES))

        if data.get("negotiation_status"):
            data["negotiation_status"] = "|".join(
                [
                    "#".join(
                        [
                            str(x.get("date", "")),
                            "current" if x.get("current") else "",
                            current_negotiation_status_map[x.get("value")],
                        ]
                    )
                    for x in data["negotiation_status"]
                ]
            )

        if data.get("implementation_status"):
            data["implementation_status"] = "|".join(
                [
                    "#".join(
                        [
                            str(x.get("date", "")),
                            "current" if x.get("current") else "",
                            current_implementation_status_map[x.get("value")],
                        ]
                    )
                    for x in data["implementation_status"]
                ]
            )

        if data.get("purchase_price"):
            data["purchase_price"] = int(data["purchase_price"])
        if data.get("purchase_price_currency"):
            data["purchase_price_currency"] = mchoices.get("currency")[
                data["purchase_price_currency"]
            ]
        if data.get("purchase_price_type"):
            data["purchase_price_type"] = dict(Deal.HA_AREA_CHOICES)[
                data["purchase_price_type"]
            ]
        if data.get("purchase_price_area"):
            data["purchase_price_area"] = int(data["purchase_price_area"])

        if data.get("annual_leasing_fee"):
            data["annual_leasing_fee"] = int(data["annual_leasing_fee"])
        if data.get("annual_leasing_fee_currency"):
            data["annual_leasing_fee_currency"] = mchoices.get("currency")[
                data["annual_leasing_fee_currency"]
            ]
        if data.get("annual_leasing_fee_type"):
            data["annual_leasing_fee_type"] = dict(Deal.HA_AREA_CHOICES)[
                data["annual_leasing_fee_type"]
            ]
        if data.get("annual_leasing_fee_area"):
            data["annual_leasing_fee_area"] = int(data["annual_leasing_fee_area"])

        bool_cast(data, "contract_farming")
        flatten_date_current_value(data, "total_jobs_current")
        flatten_date_current_value(data, "total_jobs_current_employees")
        flatten_date_current_value(data, "total_jobs_current_daily_workers")
        flatten_date_current_value(data, "foreign_jobs_current")
        flatten_date_current_value(data, "foreign_jobs_current_employees")
        flatten_date_current_value(data, "foreign_jobs_current_daily_workers")
        flatten_date_current_value(data, "domestic_jobs_current")
        flatten_date_current_value(data, "domestic_jobs_current_employees")
        flatten_date_current_value(data, "domestic_jobs_current_daily_workers")

        flatten_array_choices(
            data, "recognition_status", dict(Deal.RECOGNITION_STATUS_CHOICES)
        )
        single_choice(
            data, "community_consultation", dict(Deal.COMMUNITY_CONSULTATION_CHOICES)
        )
        single_choice(data, "community_reaction", dict(Deal.COMMUNITY_REACTION_CHOICES))

        if data.get("involved_actors"):
            data["involved_actors"] = "|".join(
                [
                    "#".join(
                        [
                            x.get("value", "") or "",
                            dict(Deal.ACTOR_MAP)[x["role"]] if x.get("role") else "",
                        ]
                    )
                    for x in data["involved_actors"]
                ]
            )

        bool_cast(data, "on_the_lease")
        flatten_date_current_value(data, "on_the_lease_area")
        flatten_date_current_value(data, "on_the_lease_farmers")
        flatten_date_current_value(data, "on_the_lease_households")
        bool_cast(data, "off_the_lease")
        flatten_date_current_value(data, "off_the_lease_area")
        flatten_date_current_value(data, "off_the_lease_farmers")
        flatten_date_current_value(data, "off_the_lease_households")
        bool_cast(data, "total_jobs_created")
        bool_cast(data, "foreign_jobs_created")
        bool_cast(data, "domestic_jobs_created")

        if data.get("name_of_community"):
            data["name_of_community"] = "".join(
                [f"{x}#" for x in data["name_of_community"]]
            )
        if data.get("name_of_indigenous_people"):
            data["name_of_indigenous_people"] = "".join(
                [f"{x}#" for x in data["name_of_indigenous_people"]]
            )

        bool_cast(data, "land_conflicts")
        bool_cast(data, "displacement_of_people")

        flatten_array_choices(
            data, "negative_impacts", dict(Deal.NEGATIVE_IMPACTS_CHOICES)
        )
        flatten_array_choices(data, "promised_benefits", dict(Deal.BENEFITS_CHOICES))
        flatten_array_choices(
            data, "materialized_benefits", dict(Deal.BENEFITS_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_owner", dict(Deal.FORMER_LAND_OWNER_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_use", dict(Deal.FORMER_LAND_USE_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_cover", dict(Deal.FORMER_LAND_COVER_CHOICES)
        )

        bool_cast(data, "has_domestic_use")
        bool_cast(data, "has_export")

        bool_cast(data, "in_country_processing")
        bool_cast(data, "water_extraction_envisaged")
        bool_cast(data, "use_of_irrigation_infrastructure")
        bool_cast(data, "fully_updated")
        bool_cast(data, "confidential")

        for country in ["export_country1", "export_country2", "export_country3"]:
            if data.get(country):
                data[country] = mchoices.get("country")[data[country]]

        for produce_type in ["crops", "animals", "resources"]:
            if data.get(produce_type) is not None:
                data[produce_type] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                dat.get("hectares") or "",
                                dat.get("tons") or "",
                                dat.get("percent") or "",
                                ", ".join(
                                    [
                                        mchoices.get(produce_type).get(x, x)
                                        for x in dat.get("value")
                                    ]
                                ),
                            ]
                        )
                        for dat in data[produce_type]
                    ]
                )

        # broken because sync not correct yet
        for produce_type in ["contract_farming_crops", "contract_farming_animals"]:
            if data.get(produce_type) is not None:
                data[produce_type] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                dat.get("hectares") or "",
                                ", ".join(
                                    [
                                        mchoices.get(produce_type[17:]).get(x, x)
                                        for x in dat.get("value")
                                    ]
                                ),
                            ]
                        )
                        for dat in data[produce_type]
                    ]
                )

        """missing
            source_of_water_extraction
            water_extraction_amount
            use_of_irrigation_infrastructure
            water_footprint
            gender_related_information
            vggt_applied
            prai_applied
            confidential_reason
        """

        return [
            "" if field not in data else data[field] for field in deal_fields.keys()
        ]

    @staticmethod
    def location_download_format(data):
        if data.get("point"):
            data["point"] = f"{data['point'].y},{data['point'].x}"
        if data.get("level_of_accuracy"):
            data["level_of_accuracy"] = dict(Location.ACCURACY_CHOICES)[
                data["level_of_accuracy"]
            ]
        return [
            "" if field not in data else data[field] for field in location_fields.keys()
        ]

    @staticmethod
    def contracts_download_format(data):
        return [
            "" if field not in data else data[field] for field in contract_fields.keys()
        ]

    @staticmethod
    def datasource_download_format(data):
        if data.get("type"):
            data["type"] = dict(DataSource.TYPE_CHOICES)[data["type"]]
        return [
            "" if field not in data else data[field]
            for field in datasource_fields.keys()
        ]

    @staticmethod
    def investor_download_format(data):
        if "country" in data:
            data["country__name"] = data["country"]["name"]

        row = []
        for field in investor_fields:
            if field not in data:
                # empty fields
                data[field] = ""
            elif field in investor_choices_fields:
                # fields with choices
                data[field] = str(investor_choices_fields[field][data[field]])
            row.append(data[field])
        return row

    @staticmethod
    def involvement_download_format(data):
        data["venture__name"] = data["venture"]["name"]
        data["investor__name"] = data["investor"]["name"]

        if "investment_type" in data:
            data["investment_type"] = "|".join(
                arrayfield_choices_display(
                    data["investment_type"],
                    InvestorVentureInvolvement._meta.get_field(
                        "investment_type"
                    ).choices,
                )
            )
        row = []
        for field in involvement_fields:
            if field not in data:
                # empty fields
                data[field] = ""
            elif field in involvement_choices_fields:
                # fields with choices
                data[field] = str(involvement_choices_fields[field][data[field]])
            row.append(data[field])
        return row
