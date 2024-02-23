import zipfile
from datetime import datetime
from io import BytesIO

import unicodecsv as csv
from django.contrib.postgres.expressions import ArraySubquery
from django.db.models import QuerySet, Func, Value, CharField, Case, When, OuterRef
from django.db.models.functions import Concat, JSONObject
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

from apps.landmatrix.involvement_network import InvolvementNetwork
from apps.landmatrix.models import choices
from apps.landmatrix.models.country import Country
from apps.landmatrix.models.currency import Currency
from apps.landmatrix.models.new import (
    DealHull,
    Location,
    Contract,
    DealDataSource,
    DealVersion,
    InvestorHull,
    Involvement,
    DealTopInvestors2,
)
from apps.landmatrix.utils import parse_filters

deal_fields = {
    "deal_id": "Deal ID",
    "is_public": "Is public",
    "transnational": "Deal scope",
    "deal_size": "Deal size",
    "deal__country": "Target country",
    "current_contract_size": "Current size under contract",
    "current_production_size": "Current size in operation (production)",
    "current_negotiation_status": "Current negotiation status",
    "current_implementation_status": "Current implementation status",
    "deal__fully_updated_at": "Fully updated",
    "top_investors": "Top parent companies",
    "intended_size": "Intended size (in ha)",  # TODO LATER 100% Compat hacks
    "contract_size": "Size under contract (leased or purchased area, in ha)",
    "production_size": "Size in operation (production, in ha)",
    "land_area_comment": "Comment on land area",
    "intention_of_investment": "Intention of investment",
    "intention_of_investment_comment": "Comment on intention of investment",
    "nature_of_deal": "Nature of the deal",
    "nature_of_deal_comment": "Comment on nature of the deal",
    "negotiation_status": "Negotiation status",
    "negotiation_status_comment": "Comment on negotiation status",
    "implementation_status": "Implementation status",
    "implementation_status_comment": "Comment on implementation status",
    "purchase_price": "Purchase price",
    "purchase_price_currency": "Purchase price currency",
    "purchase_price_type": "Purchase price area type",
    "purchase_price_area": "Purchase price area",
    "purchase_price_comment": "Comment on purchase price",
    "annual_leasing_fee": "Annual leasing fee",
    "annual_leasing_fee_currency": "Annual leasing fee currency",
    "annual_leasing_fee_type": "Annual leasing fee type",
    "annual_leasing_fee_area": "Annual leasing fee area",
    "annual_leasing_fee_comment": "Comment on leasing fees",
    "contract_farming": "Contract farming",
    "on_the_lease_state": "On leased / purchased",
    "on_the_lease": "On leased area/farmers/households",
    "off_the_lease_state": "Not on leased / purchased (out-grower)",
    "off_the_lease": "Not on leased area/farmers/households (out-grower)",
    "contract_farming_comment": "Comment on contract farming",
    "total_jobs_created": "Jobs created (total)",
    "total_jobs_planned": "Planned number of jobs (total)",
    "total_jobs_planned_employees": "Planned employees (total)",
    "total_jobs_planned_daily_workers": "Planned daily/seasonal workers (total)",
    "total_jobs_current": "Current total number of jobs/employees/ daily/seasonal workers",
    "total_jobs_created_comment": "Comment on jobs created (total)",
    "foreign_jobs_created": "Jobs created (foreign)",
    "foreign_jobs_planned": "Planned number of jobs (foreign)",
    "foreign_jobs_planned_employees": "Planned employees (foreign)",
    "foreign_jobs_planned_daily_workers": "Planned daily/seasonal workers (foreign)",
    "foreign_jobs_current": "Current foreign number of jobs/employees/ daily/seasonal workers",
    "foreign_jobs_created_comment": "Comment on jobs created (foreign)",
    "domestic_jobs_created": "Jobs created (domestic)",
    "domestic_jobs_planned": "Planned number of jobs (domestic)",
    "domestic_jobs_planned_employees": "Planned employees (domestic)",
    "domestic_jobs_planned_daily_workers": "Planned daily/seasonal workers (domestic)",
    "domestic_jobs_current": "Current domestic number of jobs/employees/ daily/seasonal workers",
    "domestic_jobs_created_comment": "Comment on jobs created (domestic)",
    "involved_actors": "Actors involved in the negotiation / admission process",
    "project_name": "Name of investment project",
    "investment_chain_comment": "Comment on investment chain",
    "operating_company__active_version__investor_id": "Operating company: Investor ID",
    "operating_company__active_version__name": "Operating company: Name",
    "operating_company__active_version__country__name": "Operating company: Country of registration/origin",
    "operating_company__active_version__classification": "Operating company: Classification",
    "operating_company__active_version__homepage": "Operating company: Investor homepage",
    "operating_company__active_version__opencorporates": "Operating company: Opencorporates link",
    "operating_company__active_version__comment": "Operating company: Comment",
    "name_of_community": "Name of community",
    "name_of_indigenous_people": "Name of indigenous people",
    "people_affected_comment": "Comment on communities / indigenous peoples affected",
    "recognition_status": "Recognition status of community land tenure",
    "recognition_status_comment": "Comment on recognitions status of community land tenure",
    "community_consultation": "Community consultation",
    "community_consultation_comment": "Comment on consultation of local community",
    "community_reaction": "Community reaction",
    "community_reaction_comment": "Comment on community reaction",
    "land_conflicts": "Presence of land conflicts",
    "land_conflicts_comment": "Comment on presence of land conflicts",
    "displacement_of_people": "Displacement of people",
    "displaced_people": "Number of people actually displaced",
    "displaced_households": "Number of households actually displaced",
    "displaced_people_from_community_land": "Number of people displaced out of their community land",
    "displaced_people_within_community_land": "Number of people displaced staying on community land",
    "displaced_households_from_fields": 'Number of households displaced "only" from their agricultural fields',
    "displaced_people_on_completion": "Number of people facing displacement once project is fully implemented",
    "displacement_of_people_comment": "Comment on displacement of people",
    "negative_impacts": "Negative impacts for local communities",
    "negative_impacts_comment": "Comment on negative impacts for local communities",
    "promised_compensation": "Promised compensation (e.g. for damages or resettlements)",
    "received_compensation": "Received compensation (e.g. for damages or resettlements)",
    "promised_benefits": "Promised benefits for local communities",
    "promised_benefits_comment": "Comment on promised benefits for local communities",
    "materialized_benefits": "Materialized benefits for local communities",
    "materialized_benefits_comment": "Comment on materialized benefits for local communities",
    "presence_of_organizations": "Presence of organizations and actions taken (e.g. farmer organizations, NGOs, etc.)",
    "former_land_owner": "Former land owner",
    "former_land_owner_comment": "Comment on former land owner",
    "former_land_use": "Former land use",
    "former_land_use_comment": "Comment on former land use",
    "former_land_cover": "Former land cover",
    "former_land_cover_comment": "Comment on former land cover",
    "crops": "Crops area/yield/export",
    "crops_comment": "Comment on crops",
    "animals": "Livestock area/yield/export",
    "animals_comment": "Comment on livestock",
    "mineral_resources": "Mineral resources area/yield/export",
    "mineral_resources_comment": "Comment on mineral resources",
    "contract_farming_crops": "Contract farming crops",
    "contract_farming_crops_comment": "Comment on contract farming crops",
    "contract_farming_animals": "Contract farming livestock",
    "contract_farming_animals_comment": "Comment on contract farming livestock",
    "electricity_generation": "Electricity generation",
    "electricity_generation_comment": "Comment on electricity generation",
    "carbon_sequestration": "Carbon sequestration",
    "carbon_sequestration_comment": "Comment on carbon sequestration",
    "has_domestic_use": "Has domestic use",
    "domestic_use": "Domestic use",
    "has_export": "Has export",
    "export": "Export",
    "export_country1": "Country 1",
    "export_country1_ratio": "Country 1 ratio",
    "export_country2": "Country 2",
    "export_country2_ratio": "Country 2 ratio",
    "export_country3": "Country 3",
    "export_country3_ratio": "Country 3 ratio",
    "use_of_produce_comment": "Comment on use of produce",
    "in_country_processing": "In country processing of produce",
    "in_country_processing_comment": "Comment on in country processing of produce",
    "in_country_processing_facilities": "Processing facilities / production infrastructure of the project (e.g. oil mill, ethanol distillery, biomass power plant etc.)",
    "in_country_end_products": "In-country end products of the project",
    "water_extraction_envisaged": "Water extraction envisaged",
    "water_extraction_envisaged_comment": "Comment on water extraction envisaged",
    "source_of_water_extraction": "Source of water extraction",
    "source_of_water_extraction_comment": "Comment on source of water extraction",
    "how_much_do_investors_pay_comment": "Comment on how much do investors pay for water",
    "water_extraction_amount": "Water extraction amount",
    "water_extraction_amount_comment": "Comment on how much water is extracted",
    "use_of_irrigation_infrastructure": "Use of irrigation infrastructure",
    "use_of_irrigation_infrastructure_comment": "Comment on use of irrigation infrastructure",
    "water_footprint": "Water footprint of the investment project",
    "gender_related_information": "Comment on gender-related info",
    "overall_comment": "Overall comment",
    "deal__confidential": "Not public",
    "deal__confidential_comment": "Comment on not public",
}

deal_headers = deal_fields.values()
location_headers = [
    "ID",
    "Deal ID",
    "Spatial accuracy level",
    "Location",
    "Point",
    "Facility name",
    "Location description",
    "Comment on location",
]
contract_headers = [
    "ID",
    "Deal ID",
    "Contract number",
    "Contract date",
    "Contract expiration date",
    "Duration of the agreement",
    "Comment on contract",
]
datasource_headers = [
    "ID",
    "Deal ID",
    "Data source type",
    "URL",
    "File",
    "Publication title",
    "Date",
    "Name",
    "Organisation",
    "Email",
    "Phone",
    "Open Contracting ID",
    "Comment on data source",
]

negotiation_status_choices = dict(choices.NEGOTIATION_STATUS_CHOICES) | {None: "None"}
implementation_status_choices = dict(choices.IMPLEMENTATION_STATUS_CHOICES) | {
    None: "None"
}
intention_of_investment_choices = dict(choices.INTENTION_OF_INVESTMENT_CHOICES) | {
    None: "None"
}
classification_choices = dict(choices.INVESTOR_CLASSIFICATION_CHOICES)
produce_choices = {
    "crops": {item["value"]: item["label"] for item in choices.CROPS_ITEMS},
    "contract_farming_crops": {
        item["value"]: item["label"] for item in choices.CROPS_ITEMS
    },
    "animals": {item["value"]: item["label"] for item in choices.ANIMALS_ITEMS},
    "contract_farming_animals": {
        item["value"]: item["label"] for item in choices.ANIMALS_ITEMS
    },
    "mineral_resources": {
        item["value"]: item["label"] for item in choices.MINERALS_ITEMS
    },
}


class Choices:
    """need to generate the choices on running, otherwise DB errors"""

    choices = {}

    def get(self, name):
        if not self.choices.get(name):
            if name == "currency":
                self.choices[name] = {
                    x.id: f"{x.name} ({x.symbol})" if x.symbol else x.name
                    for x in Currency.objects.all()
                }
            if name == "country":
                self.choices[name] = dict(Country.objects.values_list("id", "name"))
        return self.choices[name]


mchoices = Choices()

investor_headers = [
    "Investor ID",
    "Name",
    "Country of registration/origin",
    "Classification",
    "Investor homepage",
    "Opencorporates link",
    "Comment",
    "Action comment",
]
investor_fields = [
    "id",
    "name",
    "country__name",
    "classification",
    "homepage",
    "opencorporates",
    "comment",
]

involvement_headers = [
    "Involvement ID",
    "Investor ID Downstream",
    "Investor Name Downstream",
    "Investor ID Upstream",
    "Investor Name Upstream",
    "Relation type",
    "Investment type",
    "Ownership share",
    "Loan amount",
    "Loan currency",
    "Loan date",
    "Comment",
]
involvement_fields = [
    "id",
    "child_investor_id",
    "child_investor__active_version__name",
    "parent_investor_id",
    "parent_investor__active_version__name",
    "role",
    "investment_type",
    "percentage",
    "loans_amount",
    "loans_currency",
    "loans_date",
    "comment",
]


def flatten_date_current_value(data, field, fieldname) -> None:
    if not data.get(field):
        data[field] = ""
        return
    data[field] = "|".join(
        [
            "#".join(
                [
                    x["date"] or "",
                    "current" if x.get("current") else "",
                    (
                        x[fieldname]
                        if isinstance(x[fieldname], str)
                        else str(round(x[fieldname], 2))
                    ),
                ]
            )
            for x in data[field]
            if x.get(fieldname) is not None
        ]
    )


def flatten_array_choices(data, field, choics) -> None:
    if not data.get(field):
        data[field] = ""
        return

    matches = []
    for x in data[field]:
        if choice := choics.get(x):
            matches += [str(choice)]
    data[field] = "|".join(matches)


def single_choice(data, field, choics) -> None:
    if not data.get(field):
        return
    data[field] = choics[data[field]]


def bool_cast(data, field) -> None:
    if data.get(field) is None:
        return
    data[field] = "Yes" if data[field] else "No"


def deal_qs_to_values(qs: QuerySet[DealVersion]):
    return (
        qs.values(
            "deal_id",
            "is_public",
            "transnational",
            "deal_size",
            "current_contract_size",
            "current_production_size",
            "current_negotiation_status",
            "current_implementation_status",
            "intended_size",
            "contract_size",
            "production_size",
            "land_area_comment",
            "intention_of_investment",
            "intention_of_investment_comment",
            "nature_of_deal",
            "nature_of_deal_comment",
            "negotiation_status",
            "negotiation_status_comment",
            "implementation_status",
            "implementation_status_comment",
            "purchase_price",
            "purchase_price_currency",
            "purchase_price_type",
            "purchase_price_area",
            "purchase_price_comment",
            "annual_leasing_fee",
            "annual_leasing_fee_currency",
            "annual_leasing_fee_type",
            "annual_leasing_fee_area",
            "annual_leasing_fee_comment",
            "contract_farming",
            "on_the_lease_state",
            "on_the_lease",
            "off_the_lease_state",
            "off_the_lease",
            "contract_farming_comment",
            "total_jobs_created",
            "total_jobs_planned",
            "total_jobs_planned_employees",
            "total_jobs_planned_daily_workers",
            "total_jobs_current",
            "total_jobs_created_comment",
            "foreign_jobs_created",
            "foreign_jobs_planned",
            "foreign_jobs_planned_employees",
            "foreign_jobs_planned_daily_workers",
            "foreign_jobs_current",
            "foreign_jobs_created_comment",
            "domestic_jobs_created",
            "domestic_jobs_planned",
            "domestic_jobs_planned_employees",
            "domestic_jobs_planned_daily_workers",
            "domestic_jobs_current",
            "domestic_jobs_created_comment",
            "involved_actors",
            "project_name",
            "investment_chain_comment",
            "name_of_community",
            "name_of_indigenous_people",
            "people_affected_comment",
            "recognition_status",
            "recognition_status_comment",
            "community_consultation",
            "community_consultation_comment",
            "community_reaction",
            "community_reaction_comment",
            "land_conflicts",
            "land_conflicts_comment",
            "displacement_of_people",
            "displaced_people",
            "displaced_households",
            "displaced_people_from_community_land",
            "displaced_people_within_community_land",
            "displaced_households_from_fields",
            "displaced_people_on_completion",
            "displacement_of_people_comment",
            "negative_impacts",
            "negative_impacts_comment",
            "promised_compensation",
            "received_compensation",
            "promised_benefits",
            "promised_benefits_comment",
            "materialized_benefits",
            "materialized_benefits_comment",
            "presence_of_organizations",
            "former_land_owner",
            "former_land_owner_comment",
            "former_land_use",
            "former_land_use_comment",
            "former_land_cover",
            "former_land_cover_comment",
            "crops",
            "crops_comment",
            "animals",
            "animals_comment",
            "mineral_resources",
            "mineral_resources_comment",
            "contract_farming_crops",
            "contract_farming_crops_comment",
            "contract_farming_animals",
            "contract_farming_animals_comment",
            "electricity_generation",
            "electricity_generation_comment",
            "carbon_sequestration",
            "carbon_sequestration_comment",
            "has_domestic_use",
            "domestic_use",
            "has_export",
            "export",
            "export_country1",
            "export_country1_ratio",
            "export_country2",
            "export_country2_ratio",
            "export_country3",
            "export_country3_ratio",
            "use_of_produce_comment",
            "in_country_processing",
            "in_country_processing_comment",
            "in_country_processing_facilities",
            "in_country_end_products",
            "water_extraction_envisaged",
            "water_extraction_envisaged_comment",
            "source_of_water_extraction",
            "source_of_water_extraction_comment",
            "how_much_do_investors_pay_comment",
            "water_extraction_amount",
            "water_extraction_amount_comment",
            "use_of_irrigation_infrastructure",
            "use_of_irrigation_infrastructure_comment",
            "water_footprint",
            "gender_related_information",
            "overall_comment",
            "deal__confidential",
            "deal__confidential_comment",
            "deal__country",
            "deal__fully_updated_at",
        )
        .annotate(
            operating_company=Case(
                When(operating_company__active_version_id=None, then=None),
                default=JSONObject(
                    investor_id="operating_company__active_version__investor_id",
                    name="operating_company__active_version__name",
                    country__name="operating_company__active_version__country__name",
                    classification="operating_company__active_version__classification",
                    homepage="operating_company__active_version__homepage",
                    opencorporates="operating_company__active_version__opencorporates",
                    comment="operating_company__active_version__comment",
                ),
            )
        )
        .annotate(
            top_investors=(
                ArraySubquery(
                    DealTopInvestors2.objects.exclude(investorhull__active_version=None)
                    .filter(investorhull__deleted=False)
                    .filter(dealversion2_id=OuterRef("id"))
                    .order_by("-id")
                    .annotate(
                        active_version=JSONObject(
                            name="investorhull__active_version__name",
                            country_name="investorhull__active_version__country__name",
                        )
                    )
                    .values(
                        json=JSONObject(
                            id="investorhull_id", active_version="active_version"
                        )
                    )
                )
            )
        )
    )


class DataDownload:
    def __init__(self, request):
        self.request = request
        self.user = request.user
        deal_id = self.request.GET.get("deal_id")
        self.subset = self.request.GET.get("subset", "PUBLIC")
        self.return_format = self.request.GET.get("format", "html")

        if deal_id:
            self._single_deal(deal_id)
        else:
            self._multiple_deals(self.request)

    def _single_deal(self, deal_id):
        dealhull = DealHull.objects.visible(self.user, self.subset).get(id=deal_id)
        dealversions: QuerySet[DealVersion] = DealVersion.objects.filter(
            id=dealhull.active_version_id
        )

        version = dealversions[0]

        self.deals = [
            self.deal_download_format(d) for d in deal_qs_to_values(dealversions)
        ]

        self.locations = self.location_download_format(version.locations.all())
        self.contracts = self.contracts_download_format(version.contracts.all())
        self.datasources = self.datasource_download_format(version.datasources.all())

        self.investors = []
        self.involvements = []
        if version.operating_company_id:
            (
                all_investors,
                all_involvements,
                edges,
                min_depth,
            ) = InvolvementNetwork().get_network(
                version.operating_company_id,
                depth=1,
                show_ventures=False,
            )

            self.investors = self.investor_download_format(all_investors)
            self.involvements = self.involvement_download_format(all_involvements)

        self.filename = f"deal_{deal_id}"

    def _multiple_deals(self, filtersx):
        print(filtersx)
        dealhulls = DealHull.objects.visible(self.user, subset=self.subset)
        if filtersx:
            dealhulls = dealhulls.filter(parse_filters(filtersx))

        qs: QuerySet[DealVersion] = DealVersion.objects.filter(
            id__in=dealhulls.values_list("active_version_id", flat=True)
        ).order_by("deal_id")

        self.deals = [self.deal_download_format(d) for d in deal_qs_to_values(qs)]

        self.locations = []
        for d in qs:
            self.locations += self.location_download_format(d.locations.all())

        self.contracts = []
        for d in qs:
            self.contracts += self.contracts_download_format(d.contracts.all())

        self.datasources = []
        for d in qs:
            self.datasources += self.datasource_download_format(d.datasources.all())

        self.investors = self.investor_download_format(
            InvestorHull.objects.visible(self.user).order_by("id")
        )

        involvements = Involvement.objects.visible(self.user).order_by("id")
        self.involvements = self.involvement_download_format(involvements)

        self.filename = "export"

    def get_response(self):
        if self.return_format == "xlsx":
            return self.xlsx()
        if self.return_format == "csv":
            return self.csv()
        return self.html()

    def html(self):
        ctx = {
            "deal_headers": deal_headers,
            "deals": self.deals,
            "location_headers": location_headers,
            "locations": self.locations,
            "contract_headers": contract_headers,
            "contracts": self.contracts,
            "datasource_headers": datasource_headers,
            "datasources": self.datasources,
            "investor_headers": investor_headers,
            "investors": self.investors,
            "involvement_headers": involvement_headers,
            "involvements": self.involvements,
        }
        return render(self.request, "landmatrix/export_table.html", ctx)

    def xlsx(self):
        response = HttpResponse(
            headers={
                "Content-Type": "application/ms-excel",
                "Content-Disposition": f"attachment; filename={self.filename}.xlsx",
            }
        )
        wb = Workbook(write_only=True)
        # wb = Workbook()

        # Deals tab
        ws_deals = wb.create_sheet(title="Deals")
        ws_deals.append(list(deal_headers))
        for item in self.deals:
            try:
                ws_deals.append(item)
            except IllegalCharacterError:  # pragma: no cover
                ws_deals.append(
                    [str(i).encode("unicode_escape").decode("utf-8") for i in item]
                )

        # # Locations tab
        ws = wb.create_sheet(title="Locations")
        ws.append(location_headers)
        [ws.append(item) for item in self.locations]

        # # Contracts tab
        ws = wb.create_sheet(title="Contracts")
        ws.append(contract_headers)
        [ws.append(item) for item in self.contracts]

        # # DataSources tab
        ws = wb.create_sheet(title="Data sources")
        ws.append(datasource_headers)
        [ws.append(item) for item in self.datasources]

        # Involvements tab
        ws_involvements = wb.create_sheet(title="Involvements")
        ws_involvements.append(involvement_headers)
        for item in self.involvements:
            ws_involvements.append(item)

        # Investors tab
        ws_investors = wb.create_sheet(title="Investors")
        ws_investors.append(investor_headers)
        for item in self.investors:
            ws_investors.append(item)

        wb.save(response)
        return response

    @staticmethod
    def _csv_writer(data):
        # Deals CSV
        file = BytesIO()
        writer = csv.writer(file, delimiter=";")  # encoding='cp1252'
        for item in data:
            writer.writerow(item)
        file.seek(0)
        return file.getvalue()

    def csv(self):
        result = BytesIO()
        zip_file = zipfile.ZipFile(result, "w")

        # Deals CSV
        deals_data = [list(deal_headers)] + self.deals
        zip_file.writestr("deals.csv", self._csv_writer(deals_data))

        # Locations CSV
        zip_file.writestr(
            "locations.csv",
            self._csv_writer([location_headers] + self.locations),
        )
        # Contracts CSV
        zip_file.writestr(
            "contracts.csv",
            self._csv_writer([contract_headers] + self.contracts),
        )
        # Datasources CSV
        zip_file.writestr(
            "datasources.csv",
            self._csv_writer([datasource_headers] + self.datasources),
        )

        # Involvements CSV
        involvements_data = [involvement_headers] + self.involvements
        zip_file.writestr("involvements.csv", self._csv_writer(involvements_data))

        # Investors CSV
        investors_data = [investor_headers] + self.investors
        zip_file.writestr("investors.csv", self._csv_writer(investors_data))

        zip_file.close()

        return HttpResponse(
            result.getvalue(),
            headers={
                "Content-Type": "application/x-zip-compressed",
                "Content-Disposition": f"attachment; filename={self.filename}.zip",
            },
        )

    @staticmethod
    def deal_download_format(data):
        bool_cast(data, "is_public")
        if "transnational" in data:
            data["transnational"] = (
                "transnational" if data["transnational"] else "domestic"
            )

        # flatten top investors
        data["top_investors"] = "|".join(
            [
                "#".join(
                    [
                        ti["active_version"]
                        .get("name", "")
                        .replace("#", "")
                        .replace("\n", "")
                        .strip(),
                        str(ti["id"]),
                        ti["active_version"].get("country_name", "") or "",
                    ]
                )
                for ti in sorted(data["top_investors"], key=lambda x: x["id"])
                if ti["id"]
            ]
        )

        # map operating company fields
        if oc := data["operating_company"]:
            data["operating_company__active_version__investor_id"] = oc["investor_id"]
            data["operating_company__active_version__name"] = oc["name"]

            data["operating_company__active_version__country__name"] = oc[
                "country__name"
            ]
            data["operating_company__active_version__classification"] = (
                classification_choices.get(oc["classification"], "")
            )

            data["operating_company__active_version__homepage"] = oc["homepage"]
            data["operating_company__active_version__opencorporates"] = oc[
                "opencorporates"
            ]
            data["operating_company__active_version__comment"] = oc["comment"]

        # TODO LATER 100% Compat hacks
        data["deal_size"] = f'{data["deal_size"]:.2f}' if data["deal_size"] else 0.0
        data["current_contract_size"] = (
            f'{data["current_contract_size"]:.2f}'
            if data["current_contract_size"]
            else 0.0
        )
        data["current_production_size"] = (
            f'{data["current_production_size"]:.2f}'
            if data["current_production_size"]
            else 0.0
        )
        data["intended_size"] = (
            f'{data["intended_size"]:.2f}' if data["intended_size"] else ""
        )

        imp_stat = data.get("current_implementation_status")
        data["current_implementation_status"] = implementation_status_choices.get(
            imp_stat, imp_stat
        )

        neg_stat = data.get("current_negotiation_status")
        data["current_negotiation_status"] = negotiation_status_choices.get(
            neg_stat, neg_stat
        )

        fully_updated_at = data["deal__fully_updated_at"]
        if fully_updated_at:
            data["deal__fully_updated_at"] = fully_updated_at.isoformat()

        flatten_date_current_value(data, "contract_size", "area")
        flatten_date_current_value(data, "production_size", "area")

        data["intention_of_investment"] = "|".join(
            [
                "#".join(
                    [
                        x["date"] or "",
                        "current" if x.get("current") else "",
                        str(x.get("area", "") or ""),
                        ", ".join(
                            str(intention_of_investment_choices[y])
                            for y in x.get("choices", [])
                        ),
                    ]
                )
                for x in data["intention_of_investment"]
                if x.get("choices") is not None
            ]
        )

        flatten_array_choices(
            data, "nature_of_deal", dict(choices.NATURE_OF_DEAL_CHOICES)
        )

        data["negotiation_status"] = "|".join(
            [
                "#".join(
                    [
                        x["date"] or "",
                        "current" if x.get("current") else "",
                        str(negotiation_status_choices[x.get("choice")]),
                    ]
                )
                for x in data["negotiation_status"]
            ]
        )

        data["implementation_status"] = "|".join(
            [
                "#".join(
                    [
                        x["date"] or "",
                        "current" if x.get("current") else "",
                        str(implementation_status_choices[x.get("choice")]),
                    ]
                )
                for x in data["implementation_status"]
            ]
        )

        if data.get("purchase_price"):
            data["purchase_price"] = int(data["purchase_price"])
        if data.get("purchase_price_currency"):
            data["purchase_price_currency"] = mchoices.get("currency")[
                data["purchase_price_currency"]
            ]
        if data.get("purchase_price_type"):
            data["purchase_price_type"] = dict(choices.HA_AREA_CHOICES)[
                data["purchase_price_type"]
            ]
        if data.get("purchase_price_area"):
            data["purchase_price_area"] = f'{data["purchase_price_area"]:.2f}'
        if data.get("annual_leasing_fee"):
            data["annual_leasing_fee"] = int(data["annual_leasing_fee"])
        if data.get("annual_leasing_fee_currency"):
            data["annual_leasing_fee_currency"] = mchoices.get("currency")[
                data["annual_leasing_fee_currency"]
            ]
        if data.get("annual_leasing_fee_type"):
            data["annual_leasing_fee_type"] = dict(choices.HA_AREA_CHOICES)[
                data["annual_leasing_fee_type"]
            ]
        if data.get("annual_leasing_fee_area"):
            # TODO LATER 100% Compat hacks
            data["annual_leasing_fee_area"] = (
                f'{data["annual_leasing_fee_area"]:.2f}'
                if data["annual_leasing_fee_area"]
                else 0.0
            )

        bool_cast(data, "contract_farming")

        for xdings in [
            "total_jobs_current",
            "foreign_jobs_current",
            "domestic_jobs_current",
        ]:
            if data.get(xdings) is not None:
                data[xdings] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                str(dat.get("jobs") or ""),
                                str(dat.get("employees") or ""),
                                str(dat.get("workers") or ""),
                            ]
                        )
                        for dat in data[xdings]
                    ]
                )

        flatten_array_choices(
            data, "recognition_status", dict(choices.RECOGNITION_STATUS_CHOICES)
        )
        single_choice(
            data,
            "community_consultation",
            dict(choices.COMMUNITY_CONSULTATION_CHOICES),
        )
        single_choice(
            data, "community_reaction", dict(choices.COMMUNITY_REACTION_CHOICES)
        )

        # ic(data["involved_actors"])
        if data.get("involved_actors"):
            data["involved_actors"] = "|".join(
                [
                    "#".join(
                        [
                            x.get("name") or "",
                            (
                                str(dict(choices.ACTOR_MAP)[x["role"]])
                                if x.get("role")
                                else ""
                            ),
                        ]
                    )
                    for x in data["involved_actors"]
                ]
            )
        else:
            data["involved_actors"] = ""

        bool_cast(data, "on_the_lease_state")
        bool_cast(data, "off_the_lease_state")

        for xdings in ["on_the_lease", "off_the_lease"]:
            if data.get(xdings) is not None:
                data[xdings] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                str(dat.get("area") or ""),
                                str(dat.get("farmers") or ""),
                                str(dat.get("households") or ""),
                            ]
                        )
                        for dat in data[xdings]
                    ]
                )
        bool_cast(data, "total_jobs_created")
        bool_cast(data, "foreign_jobs_created")
        bool_cast(data, "domestic_jobs_created")

        if data.get("name_of_community"):
            data["name_of_community"] = "".join(
                [f"{x}#" for x in data["name_of_community"]]
            )
        if data.get("name_of_indigenous_people"):
            data["name_of_indigenous_people"] = "".join(
                [f"{x}#" for x in data["name_of_indigenous_people"]]
            )

        bool_cast(data, "land_conflicts")
        bool_cast(data, "displacement_of_people")

        flatten_array_choices(
            data, "negative_impacts", dict(choices.NEGATIVE_IMPACTS_CHOICES)
        )
        flatten_array_choices(data, "promised_benefits", dict(choices.BENEFITS_CHOICES))
        flatten_array_choices(
            data, "materialized_benefits", dict(choices.BENEFITS_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_owner", dict(choices.FORMER_LAND_OWNER_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_use", dict(choices.FORMER_LAND_USE_CHOICES)
        )
        flatten_array_choices(
            data, "former_land_cover", dict(choices.FORMER_LAND_COVER_CHOICES)
        )

        bool_cast(data, "has_domestic_use")
        bool_cast(data, "has_export")

        bool_cast(data, "in_country_processing")
        bool_cast(data, "water_extraction_envisaged")
        # bool_cast(data, "use_of_irrigation_infrastructure")
        bool_cast(data, "fully_updated")
        bool_cast(data, "confidential")

        data["deal__country"] = mchoices.get("country")[data["deal__country"]]

        for country in [
            "export_country1",
            "export_country2",
            "export_country3",
        ]:
            if data.get(country):
                data[country] = mchoices.get("country")[data[country]]

        # if (crops := data.get("crops")) is not None:
        #     data["crops_json"] = json.dumps(crops)

        for produce_type in ["crops", "animals", "mineral_resources"]:
            if data.get(produce_type) is not None:
                data[produce_type] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                str(dat.get("area") or ""),
                                str(dat.get("yield") or ""),
                                str(dat.get("export") or ""),
                                ", ".join(
                                    [
                                        str(produce_choices.get(produce_type).get(x, x))
                                        for x in dat.get("choices", [])
                                    ]
                                ),
                            ]
                        )
                        for dat in data[produce_type]
                    ]
                )

        for produce_type in ["contract_farming_crops", "contract_farming_animals"]:
            if data.get(produce_type) is not None:
                data[produce_type] = "|".join(
                    [
                        "#".join(
                            [
                                dat.get("date") or "",
                                "current" if dat.get("current") else "",
                                str(dat.get("area", "")),
                                ", ".join(
                                    [
                                        str(produce_choices.get(produce_type).get(x, x))
                                        for x in dat.get("choices", [])
                                    ]
                                ),
                            ]
                        )
                        for dat in data[produce_type]
                    ]
                )

        if (eg := data.get("electricity_generation")) is not None:
            data["electricity_generation"] = "|".join(
                [
                    "#".join(
                        [
                            dat.get("date") or "",
                            "current" if dat.get("current") else "",
                            str(dat.get("area", "")),
                            ", ".join([str(x) for x in dat.get("choices", [])]),
                            str(dat.get("export", "")),
                            str(dat.get("windfarm_count", "")),
                            str(dat.get("current_capacity", "")),
                            str(dat.get("intended_capacity", "")),
                        ]
                    )
                    for dat in eg
                ]
            )

        if (cs := data.get("carbon_sequestration")) is not None:
            data["carbon_sequestration"] = "|".join(
                [
                    "#".join(
                        [
                            dat.get("date") or "",
                            "current" if dat.get("current") else "",
                            str(dat.get("area", "")),
                            ", ".join([str(x) for x in dat.get("choices", [])]),
                            str(dat.get("projected_lifetime_sequestration", "")),
                            str(dat.get("projected_annual_sequestration", "")),
                            str(dat.get("certification_standard", "")),
                            str(dat.get("certification_standard_name", "")),
                            str(dat.get("certification_standard_comment", "")),
                        ]
                    )
                    for dat in cs
                ]
            )

        flatten_array_choices(
            data, "source_of_water_extraction", dict(choices.WATER_SOURCE_CHOICES)
        )
        bool_cast(data, "use_of_irrigation_infrastructure")

        bool_cast(data, "deal__confidential")

        xx = [
            (
                ""
                if (field not in data or data[field] is None or data[field] == [])
                else data[field]
            )
            for field in deal_fields.keys()
        ]

        return xx

    @staticmethod
    def location_download_format(locations: QuerySet[Location]) -> list[list]:
        return [
            [
                x["nid"],
                x["dealversion__deal_id"],
                (
                    choices.LOCATION_ACCURACY_MAP[x["level_of_accuracy"]]
                    if x["level_of_accuracy"]
                    else ""
                ),
                x["name"],
                x["point_lat_lng"],
                x["facility_name"],
                x["description"],
                x["comment"],
            ]
            for x in locations.annotate(
                point_lat_lng=Case(
                    When(point=None, then=Value("")),
                    default=Concat(
                        Func("point", function="ST_Y"),
                        Value(","),
                        Func("point", function="ST_X"),
                        output_field=CharField(),
                    ),
                )
            ).values(
                "nid",
                "dealversion__deal_id",
                "level_of_accuracy",
                "name",
                "point_lat_lng",
                "facility_name",
                "description",
                "comment",
            )
        ]

    @staticmethod
    def contracts_download_format(contracts: QuerySet[Contract]):
        return [
            [
                x["nid"],
                x["dealversion__deal_id"],
                x["number"],
                x["date_or_empty"],
                x["expiration_date_or_empty"],
                x["agreement_duration"],
                x["comment"],
            ]
            for x in contracts.annotate(
                date_or_empty=Case(
                    When(date=None, then=Value("")),
                    default="date",
                    output_field=CharField(),
                ),
                expiration_date_or_empty=Case(
                    When(expiration_date=None, then=Value("")),
                    default="expiration_date",
                    output_field=CharField(),
                ),
            ).values(
                "nid",
                "dealversion__deal_id",
                "number",
                "date_or_empty",
                "expiration_date_or_empty",
                "agreement_duration",
                "comment",
            )
        ]

    @staticmethod
    def datasource_download_format(datasources: QuerySet[DealDataSource]) -> list[list]:
        return [
            [
                x["nid"],
                x["dealversion__deal_id"],
                choices.DATASOURCE_TYPE_MAP[x["type"]] if x["type"] else "",
                x["url"],
                x["public_file"],
                x["publication_title"],
                x["date_or_empty"],
                x["name"],
                x["company"],
                x["email"],
                x["phone"],
                x["open_land_contracts_id"],
                x["comment"],
            ]
            for x in datasources.annotate(
                date_or_empty=Case(
                    When(date=None, then=Value("")),
                    default="date",
                    output_field=CharField(),
                )
            )
            .annotate(
                public_file=Case(
                    When(file_not_public=False, then="file"),
                    default=Value("-redacted-"),
                    output_field=CharField(),
                )
            )
            .values(
                "nid",
                "dealversion__deal_id",
                "type",
                "url",
                "public_file",
                "publication_title",
                "date_or_empty",
                "name",
                "company",
                "email",
                "phone",
                "open_land_contracts_id",
                "comment",
            )
        ]

    @staticmethod
    def investor_download_format(investors: QuerySet[InvestorHull]):
        ret = []
        for x in (
            investors.annotate(
                classification_or_empty=Case(
                    When(active_version__classification=None, then=Value("")),
                    default="active_version__classification",
                    output_field=CharField(),
                )
            )
            .annotate(
                country_or_empty=Case(
                    When(active_version__country__name=None, then=Value("")),
                    default="active_version__country__name",
                    output_field=CharField(),
                )
            )
            .values_list(
                "id",
                "active_version__name",
                "country_or_empty",
                "classification_or_empty",
                "active_version__homepage",
                "active_version__opencorporates",
                "active_version__comment",
            )
        ):
            x = list(x)
            x[3] = classification_choices.get(x[3], "")
            ret += [x]
        return ret

    @staticmethod
    def involvement_download_format(involvements: QuerySet[Involvement]) -> list[list]:
        return [
            [
                x["id"],
                x["child_investor_id"],
                x["child_investor__active_version__name"],
                x["parent_investor_id"],
                x["parent_investor__active_version__name"],
                choices.INVOLVEMENT_ROLE_DICT[x["role"]] if x["role"] else "",
                "|".join(
                    [
                        str(choices.INVESTMENT_TYPE_MAP[y])
                        for y in x["investment_type"]
                        if y
                    ]
                ),
                float(x["percentage"]) if x["percentage"] is not None else "",
                x["loans_amount"] if x["loans_amount"] is not None else "",
                x["loans_currency"] if x["loans_currency"] is not None else "",
                x["loans_date"] if x["loans_date"] is not None else "",
                x["comment"],
            ]
            for x in involvements.values(
                "id",
                "child_investor_id",
                "child_investor__active_version__name",
                "parent_investor_id",
                "parent_investor__active_version__name",
                "role",
                "investment_type",
                "percentage",
                "loans_amount",
                "loans_currency",
                "loans_date",
                "comment",
            )
        ]
